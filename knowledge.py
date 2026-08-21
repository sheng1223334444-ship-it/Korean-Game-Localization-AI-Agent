from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook


# =========================================================
# 基础路径
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
KNOWLEDGE_DIR = BASE_DIR / "knowledge"

ROLE_FILE = KNOWLEDGE_DIR / "角色名_最终版.xlsx"
UWO_FILE = KNOWLEDGE_DIR / "UWO_最终版.xlsx"
QUEST_FILE = KNOWLEDGE_DIR / "Quest_最终版.xlsx"


# =========================================================
# 通用工具
# =========================================================

def clean_value(value):
    if value is None:
        return ""
    return str(value)


def get_header_map(sheet):

    first_row = next(
        sheet.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        ),
        []
    )

    header_map = {}

    for index, header in enumerate(first_row):

        if header is None:
            continue

        header_map[
            clean_value(header).strip()
        ] = index

    return header_map


def is_hangul_char(char):

    if not char:
        return False

    return (
        "\uac00" <= char <= "\ud7a3"
        or
        "\u1100" <= char <= "\u11ff"
        or
        "\u3130" <= char <= "\u318f"
    )


# =========================================================
# 角色名库
# =========================================================

def load_role_names():

    workbook = load_workbook(
        ROLE_FILE,
        read_only=True,
        data_only=True
    )

    sheet = workbook["正式角色名"]

    headers = get_header_map(sheet)

    required_columns = [
        "韩文角色名",
        "正式中文名"
    ]

    for column in required_columns:

        if column not in headers:

            workbook.close()

            raise ValueError(
                f"角色名库缺少必要列：{column}"
            )

    korean_col = headers["韩文角色名"]
    chinese_col = headers["正式中文名"]

    role_names = {}

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        korean = clean_value(
            row[korean_col]
        ).strip()

        chinese = clean_value(
            row[chinese_col]
        ).strip()

        if korean and chinese:
            role_names[korean] = chinese

    workbook.close()

    return role_names


# =========================================================
# UWO / Quest 正式主库
# =========================================================

def load_translation_memory(
    file_path,
    source_name
):

    workbook = load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )

    sheet = workbook["正式主库"]

    headers = get_header_map(sheet)

    for column in [
        "msgid",
        "msgstr[0]"
    ]:

        if column not in headers:

            workbook.close()

            raise ValueError(
                f"{source_name}缺少必要列：{column}"
            )

    memory = defaultdict(list)

    for row_number, row in enumerate(
        sheet.iter_rows(
            min_row=2,
            values_only=True
        ),
        start=2
    ):

        msgid = clean_value(
            row[headers["msgid"]]
        )

        msgstr = clean_value(
            row[headers["msgstr[0]"]]
        )

        if not msgid:
            continue

        # 空中文不能当正式译文
        if not msgstr:
            continue

        record = {
            "来源库": source_name,
            "原表行号": row_number,
            "msgid": msgid,
            "msgstr[0]": msgstr,

            "msgctxt": "",
            "msgid_plural": "",
            "references": "",
            "Release": "",
            "Content": "",
            "script": "",
        }

        optional_columns = [
            "msgctxt",
            "msgid_plural",
            "references",
            "Release",
            "Content",
            "script",
        ]

        for column in optional_columns:

            if column in headers:

                record[column] = clean_value(
                    row[headers[column]]
                )

        # 同一句韩文多条历史记录全部保留
        memory[msgid].append(record)

    workbook.close()

    return dict(memory)


# =========================================================
# 长术语候选规则
# =========================================================

def is_term_candidate(text):

    text = clean_value(text).strip()

    if len(text) < 2:
        return False

    if len(text) > 60:
        return False

    if "\n" in text or "\r" in text:
        return False

    sentence_endings = (
        ".",
        "?",
        "!",
        "。",
        "？",
        "！"
    )

    if text.endswith(sentence_endings):
        return False

    return True


# =========================================================
# KnowledgeBase
# =========================================================

class KnowledgeBase:

    def __init__(self):

        print("正在加载角色名库...")

        self.role_names = load_role_names()


        print("正在加载 UWO 正式主库...")

        self.uwo = load_translation_memory(
            UWO_FILE,
            "UWO"
        )


        print("正在加载 Quest 正式主库...")

        self.quest = load_translation_memory(
            QUEST_FILE,
            "Quest"
        )


        print("正在建立长术语索引...")

        self.uwo_term_candidates = sorted(
            [
                term
                for term in self.uwo.keys()
                if is_term_candidate(term)
            ],
            key=len,
            reverse=True
        )

        self.quest_term_candidates = sorted(
            [
                term
                for term in self.quest.keys()
                if is_term_candidate(term)
            ],
            key=len,
            reverse=True
        )


        print("知识库加载完成。")


    # =====================================================
    # 角色名精确查询
    # =====================================================

    def search_role_exact(
        self,
        korean_name
    ):

        chinese = self.role_names.get(
            korean_name
        )

        if chinese:

            return {
                "命中": True,
                "韩文角色名": korean_name,
                "正式中文名": chinese,
                "来源库": "角色名库",
                "匹配状态": "精确匹配",
            }

        return {
            "命中": False
        }


    # =====================================================
    # 角色名句中识别
    # =====================================================

    def find_roles_in_text(
        self,
        text
    ):

        matches = []

        allowed_particles = [
            "으로부터",
            "에게서",
            "한테서",
            "이라도",
            "이라면",
            "이든지",
            "으로",
            "에서",
            "에게",
            "한테",
            "까지",
            "부터",
            "처럼",
            "보다",
            "께서",
            "이랑",
            "하고",
            "이나",
            "라도",
            "라면",
            "와",
            "과",
            "을",
            "를",
            "이",
            "가",
            "은",
            "는",
            "의",
            "에",
            "께",
            "로",
            "도",
            "만",
            "랑",
            "조차",
            "마저",
            "야",
            "아",
            "씨",
            "님",
        ]

        sorted_names = sorted(
            self.role_names.keys(),
            key=len,
            reverse=True
        )

        occupied_spans = []

        for korean_name in sorted_names:

            search_start = 0

            while True:

                start = text.find(
                    korean_name,
                    search_start
                )

                if start == -1:
                    break

                end = (
                    start
                    +
                    len(korean_name)
                )

                before = (
                    text[start - 1]
                    if start > 0
                    else ""
                )

                after_text = text[end:]

                # 前面直接连着韩文字：
                # 很可能只是另一个词内部
                if (
                    before
                    and
                    is_hangul_char(before)
                ):

                    search_start = start + 1
                    continue

                valid_after = True

                if after_text:

                    first_after = after_text[0]

                    if is_hangul_char(
                        first_after
                    ):

                        particle_match = False

                        for particle in allowed_particles:

                            if after_text.startswith(
                                particle
                            ):

                                particle_match = True
                                break

                        if not particle_match:
                            valid_after = False

                if not valid_after:

                    search_start = start + 1
                    continue

                overlap = False

                for (
                    old_start,
                    old_end
                ) in occupied_spans:

                    if (
                        start < old_end
                        and
                        end > old_start
                    ):

                        overlap = True
                        break

                if overlap:

                    search_start = start + 1
                    continue

                matches.append(
                    {
                        "韩文角色名":
                            korean_name,

                        "正式中文名":
                            self.role_names[
                                korean_name
                            ],

                        "来源库":
                            "角色名库",

                        "匹配状态":
                            "正式角色名",
                    }
                )

                occupied_spans.append(
                    (
                        start,
                        end
                    )
                )

                break

        return matches


    # =====================================================
    # 完整句精确查询
    # =====================================================

    def search_uwo_exact(
        self,
        korean_text
    ):

        return self.uwo.get(
            korean_text,
            []
        )


    def search_quest_exact(
        self,
        korean_text
    ):

        return self.quest.get(
            korean_text,
            []
        )


    # =====================================================
    # 字符区间重叠
    # =====================================================

    @staticmethod
    def spans_overlap(
        start,
        end,
        occupied_spans
    ):

        for (
            old_start,
            old_end
        ) in occupied_spans:

            if (
                start < old_end
                and
                end > old_start
            ):
                return True

        return False


    # =====================================================
    # 长术语检索
    # =====================================================

    def find_long_terms_from_memory(
        self,
        text,
        memory,
        candidates,
        source_name,
        min_length=4,
        blocked_terms=None
    ):

        if blocked_terms is None:
            blocked_terms = set()

        results = []

        occupied_spans = []

        for term in candidates:

            term = term.strip()

            if len(term) < min_length:
                continue

            # 完整句由精确查询负责
            if term == text:
                continue

            if term in blocked_terms:
                continue

            search_start = 0

            while True:

                position = text.find(
                    term,
                    search_start
                )

                if position == -1:
                    break

                start = position

                end = (
                    position
                    +
                    len(term)
                )

                if self.spans_overlap(
                    start,
                    end,
                    occupied_spans
                ):

                    search_start = (
                        position + 1
                    )

                    continue

                records = memory.get(
                    term,
                    []
                )

                if records:

                    results.append(
                        {
                            "韩文术语":
                                term,

                            "字符起点":
                                start,

                            "字符终点":
                                end,

                            "出现次数":
                                text.count(term),

                            "来源库":
                                source_name,

                            "匹配状态":
                                "句中长术语匹配",

                            "历史记录":
                                records,
                        }
                    )

                    occupied_spans.append(
                        (
                            start,
                            end
                        )
                    )

                break

        return results


    def search_uwo_long_terms(
        self,
        text,
        min_length=4
    ):

        return self.find_long_terms_from_memory(
            text=text,
            memory=self.uwo,
            candidates=
                self.uwo_term_candidates,
            source_name="UWO",
            min_length=min_length
        )


    def search_quest_long_terms(
        self,
        text,
        blocked_terms=None,
        min_length=4
    ):

        return self.find_long_terms_from_memory(
            text=text,
            memory=self.quest,
            candidates=
                self.quest_term_candidates,
            source_name="Quest",
            min_length=min_length,
            blocked_terms=
                blocked_terms
        )


    def search_long_terms(
        self,
        text,
        min_length=4
    ):

        uwo_terms = (
            self.search_uwo_long_terms(
                text=text,
                min_length=min_length
            )
        )

        uwo_term_names = {
            item["韩文术语"]
            for item in uwo_terms
        }

        quest_terms = (
            self.search_quest_long_terms(
                text=text,
                blocked_terms=
                    uwo_term_names,
                min_length=min_length
            )
        )

        return {
            "UWO":
                uwo_terms,

            "Quest":
                quest_terms,
        }


    # =====================================================
    # ★ 新增：历史上下文包含检索
    # =====================================================

    def search_contains_in_memory(
        self,
        query,
        memory,
        source_name,
        max_results=30
    ):
        """
        反向包含查询。

        例如用户输入：

        서양 건조

        知识库存在：

        서양 건조 LV 11 달성

        这不是完整句精确匹配，
        但知识库长文本包含了查询词。

        返回状态：

        历史上下文包含匹配

        注意：
        这种结果不能直接宣称
        “查询词就是正式术语”。
        """

        query = clean_value(
            query
        ).strip()

        if len(query) < 2:

            return {
                "总数": 0,
                "结果": [],
                "是否截断": False,
            }


        candidates = []


        for msgid, records in memory.items():

            # 完全一样的情况
            # 已经由精确匹配处理
            if msgid == query:
                continue


            if query in msgid:

                candidates.append(
                    {
                        "来源库":
                            source_name,

                        "查询词":
                            query,

                        "历史韩文":
                            msgid,

                        "历史记录":
                            records,

                        "匹配状态":
                            "历史上下文包含匹配",

                        "文本长度":
                            len(msgid),
                    }
                )


        # 优先显示：
        # 最接近用户查询长度的历史文本
        candidates.sort(
            key=lambda item: (
                item["文本长度"],
                item["历史韩文"]
            )
        )


        total = len(candidates)


        visible_results = (
            candidates[
                :max_results
            ]
        )


        return {
            "总数":
                total,

            "结果":
                visible_results,

            "是否截断":
                total > max_results,
        }


    # =====================================================
    # UWO包含查询
    # =====================================================

    def search_uwo_contains(
        self,
        query,
        max_results=30
    ):

        return (
            self.search_contains_in_memory(
                query=query,
                memory=self.uwo,
                source_name="UWO",
                max_results=max_results
            )
        )


    # =====================================================
    # Quest包含查询
    # =====================================================

    def search_quest_contains(
        self,
        query,
        max_results=30
    ):

        return (
            self.search_contains_in_memory(
                query=query,
                memory=self.quest,
                source_name="Quest",
                max_results=max_results
            )
        )


    # =====================================================
    # 基础精确检索
    # =====================================================

    def search_exact(
        self,
        korean_text
    ):

        return {
            "原文":
                korean_text,

            "角色名":
                self.find_roles_in_text(
                    korean_text
                ),

            "UWO":
                self.search_uwo_exact(
                    korean_text
                ),

            "Quest":
                self.search_quest_exact(
                    korean_text
                ),
        }


    # =====================================================
    # 当前综合查询
    # =====================================================

    def search(
        self,
        korean_text
    ):
        """
        当前查询顺序：

        1. 正式角色名
        2. UWO完整句
        3. Quest完整句
        4. UWO长术语
        5. Quest长术语
        6. UWO历史上下文包含
        7. Quest历史上下文包含

        后续：
        短术语
        相似匹配
        Content/script/references评分
        AI翻译
        """

        korean_text = clean_value(
            korean_text
        ).strip()


        exact = self.search_exact(
            korean_text
        )


        long_terms = self.search_long_terms(
            korean_text
        )


        uwo_contains = (
            self.search_uwo_contains(
                korean_text
            )
        )


        quest_contains = (
            self.search_quest_contains(
                korean_text
            )
        )


        return {
            "原文":
                korean_text,

            "角色名":
                exact["角色名"],

            "UWO完整句":
                exact["UWO"],

            "Quest完整句":
                exact["Quest"],

            "UWO长术语":
                long_terms["UWO"],

            "Quest长术语":
                long_terms["Quest"],

            "UWO包含匹配":
                uwo_contains,

            "Quest包含匹配":
                quest_contains,
        }


    # =====================================================
    # 统计
    # =====================================================

    def stats(self):

        uwo_records = sum(
            len(records)
            for records
            in self.uwo.values()
        )

        quest_records = sum(
            len(records)
            for records
            in self.quest.values()
        )

        return {
            "角色名数量":
                len(
                    self.role_names
                ),

            "UWO唯一韩文数量":
                len(
                    self.uwo
                ),

            "UWO正式译文记录数":
                uwo_records,

            "Quest唯一韩文数量":
                len(
                    self.quest
                ),

            "Quest正式译文记录数":
                quest_records,

            "UWO术语候选数量":
                len(
                    self.uwo_term_candidates
                ),

            "Quest术语候选数量":
                len(
                    self.quest_term_candidates
                ),
        }


# =========================================================
# 单独运行测试
# =========================================================

if __name__ == "__main__":

    kb = KnowledgeBase()

    print()

    print(
        "========== 知识库统计 =========="
    )

    statistics = kb.stats()

    for key, value in statistics.items():

        print(
            f"{key}: {value}"
        )


    # -----------------------------------------------------
    # 测试反向包含
    # -----------------------------------------------------

    print()

    print(
        "========== 包含检索测试 =========="
    )

    test_query = "서양 건조"

    print(
        f"查询：{test_query}"
    )


    uwo_result = (
        kb.search_uwo_contains(
            test_query,
            max_results=10
        )
    )


    print(
        f"UWO包含匹配总数："
        f"{uwo_result['总数']}"
    )


    for item in uwo_result["结果"]:

        print(
            " - "
            +
            item["历史韩文"]
        )


    print()

    print(
        "知识库读取与包含查询测试完成。"
    )