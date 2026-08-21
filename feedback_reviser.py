from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from xlsx_translator import (
    detect_header_row,
    get_headers,
    find_column_candidates,
    recommend_candidate,
    resolve_column,
    choose_sheet,
    choose_korean_column,
    cell_text,
    is_formula_cell,
    format_report_sheet,
    write_dict_rows,
)

from format_qa import (
    check_format,
    format_qa_summary,
)


# =========================================================
# 部门反馈修订模块
#
# 业务目标：
#
# 原翻译文件
# +
# 部门反馈 / 校验报告
# ↓
# 精确寻找对应记录
# ↓
# 只写回部门已经确认的“修订后中文”
# ↓
# 输出新的修订版Excel
#
#
# 核心原则：
#
# 1. 部门“修订后中文”是当前批次最高优先级。
# 2. 不调用AI自由润色。
# 3. 不自动修改长期角色名 / UWO / Quest知识库。
# 4. 不覆盖韩文原文。
# 5. 不删除、合并、拆分、重排行。
# 6. 原Excel所有其他工作表保留。
# 7. 重复韩文绝不批量扩散。
# 8. 同一韩文有多行时必须继续利用：
#       原表行号
#       ID
#       原中文
#       references
#       Release
#       Content
#       script
#       msgctxt
#    等上下文继续缩小。
# 9. 仍无法唯一确定 → 不写回，人工确认。
# 10. 修订后中文格式异常 → 不写回，人工确认。
# 11. 公式 / 合并单元格 → 不强制修改。
# 12. 最终输出新的xlsx，绝不覆盖原文件。
#
#
# 推荐匹配优先级：
#
# 原表行号
# ＞ 唯一ID
# ＞ ID + 韩文 / 原中文 / 上下文
# ＞ 韩文 + 原中文 + 上下文
# ＞ 唯一韩文
# ＞ 人工确认
#
#
# 输出辅助表：
#
# 部门反馈修订摘要
# 部门反馈修订明细
# 未匹配反馈
# 冲突待确认
# 格式异常_反馈修订
#
# =========================================================


GENERATED_SHEETS = [
    "部门反馈修订摘要",
    "部门反馈修订明细",
    "未匹配反馈",
    "冲突待确认",
    "格式异常_反馈修订",
]


# =========================================================
# 字段别名
# =========================================================

ID_ALIASES = [
    "id",
    "stringid",
    "string_id",
    "key",
    "文本id",
    "字符串id",
    "msgctxt",
]


ROW_ALIASES = [
    "原表行号",
    "原始行号",
    "源表行号",
    "原文件行号",
    "行号",
    "row",
    "rowindex",
    "row_index",
]


SHEET_ALIASES = [
    "原工作表",
    "原表",
    "工作表",
    "sheet",
    "sheetname",
    "sheet_name",
]


KOREAN_ALIASES = [
    "韩文原文",
    "韩文",
    "韩语原文",
    "韩语",
    "msgid",
    "korean",
    "source",
    "원문",
    "한국어",
]


ORIGINAL_CHINESE_ALIASES = [
    "原中文",
    "原译文",
    "修订前中文",
    "修改前中文",
    "现有中文",
    "中文原译",
    "原始中文",
    "msgstr[0]",
    "msgstr",
]


# =========================================================
# 允许作为“部门最终修订”的字段
# =========================================================

REVISED_FINAL_ALIASES = [
    "修订后中文",
    "修改后中文",
    "审校后中文",
    "最终中文",
    "最终译文",
    "确认中文",
    "确认译文",
    "部门修订后中文",
    "部门最终中文",
]


# =========================================================
# 这些只是建议，不应自动作为部门最终结论
# =========================================================

SUGGESTION_ALIASES = [
    "建议中文",
    "ai建议译文",
    "ai中文译文",
    "建议译文",
]


CONTEXT_ALIASES = {

    "msgctxt": [
        "msgctxt",
        "context",
        "上下文",
    ],

    "references": [
        "references",
        "reference",
        "refs",
        "引用",
    ],

    "Release": [
        "release",
        "版本",
    ],

    "Content": [
        "content",
        "内容",
    ],

    "script": [
        "script",
        "脚本",
    ],
}


# =========================================================
# 基础工具
# =========================================================

def normalize_header(value):

    if value is None:
        return ""

    return str(
        value
    ).strip().lower()


def clean_text(value):

    if value is None:
        return ""

    return str(
        value
    ).strip()


def normalized_aliases(values):

    return {
        normalize_header(
            value
        )
        for value in values
    }


# =========================================================
# 根据别名寻找列
# =========================================================

def find_alias_columns(
    worksheet,
    header_row,
    aliases,
):

    alias_set = normalized_aliases(
        aliases
    )

    result = []


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        header_value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        normalized = normalize_header(
            header_value
        )


        if normalized in alias_set:

            result.append(
                {
                    "列号":
                        column_index,

                    "列字母":
                        get_column_letter(
                            column_index
                        ),

                    "列名":
                        clean_text(
                            header_value
                        ),

                    "得分":
                        100,
                }
            )


    return result


# =========================================================
# 模糊寻找修订字段
# =========================================================

def find_revised_candidates(
    worksheet,
    header_row,
):

    final_aliases = normalized_aliases(
        REVISED_FINAL_ALIASES
    )

    suggestion_aliases = normalized_aliases(
        SUGGESTION_ALIASES
    )

    result = []


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        header_value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        header = normalize_header(
            header_value
        )


        if not header:
            continue


        score = 0
        field_type = ""


        if header in final_aliases:

            score = 120
            field_type = "部门最终修订"


        elif (
            "修订后" in header
            or
            "修改后" in header
            or
            "审校后" in header
            or
            "最终中文" in header
            or
            "最终译文" in header
        ):

            score = 100
            field_type = "部门最终修订"


        elif header in suggestion_aliases:

            score = 35
            field_type = "建议字段"


        elif (
            "建议" in header
            and
            (
                "中文" in header
                or
                "译文" in header
            )
        ):

            score = 30
            field_type = "建议字段"


        if score > 0:

            result.append(
                {
                    "列号":
                        column_index,

                    "列字母":
                        get_column_letter(
                            column_index
                        ),

                    "列名":
                        clean_text(
                            header_value
                        ),

                    "得分":
                        score,

                    "字段类型":
                        field_type,
                }
            )


    result.sort(
        key=lambda item:
            item["得分"],
        reverse=True,
    )


    return result


# =========================================================
# 推荐修订字段
# =========================================================

def recommend_revised_column(
    candidates
):

    final_candidates = [
        item
        for item in candidates
        if item.get(
            "字段类型"
        )
        ==
        "部门最终修订"
    ]


    if not final_candidates:
        return None


    if len(final_candidates) == 1:

        return final_candidates[
            0
        ]


    first = final_candidates[
        0
    ]


    second = final_candidates[
        1
    ]


    if (
        first.get(
            "得分",
            0,
        )
        -
        second.get(
            "得分",
            0,
        )
        >=
        15
    ):

        return first


    return None


# =========================================================
# 部门反馈表头检测
# =========================================================

def detect_feedback_header_row(
    worksheet,
    max_scan_rows=30,
):

    scan_end = min(
        worksheet.max_row,
        max_scan_rows,
    )


    all_known_aliases = set()


    for group in [
        ID_ALIASES,
        ROW_ALIASES,
        SHEET_ALIASES,
        KOREAN_ALIASES,
        ORIGINAL_CHINESE_ALIASES,
        REVISED_FINAL_ALIASES,
        SUGGESTION_ALIASES,
    ]:

        all_known_aliases.update(
            normalized_aliases(
                group
            )
        )


    for aliases in (
        CONTEXT_ALIASES.values()
    ):

        all_known_aliases.update(
            normalized_aliases(
                aliases
            )
        )


    best_row = 1
    best_score = -1


    for row_index in range(
        1,
        scan_end + 1,
    ):

        score = 0
        non_empty = 0


        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):

            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value


            if value is None:
                continue


            text = clean_text(
                value
            )


            if not text:
                continue


            non_empty += 1


            header = normalize_header(
                text
            )


            if header in all_known_aliases:

                score += 20


            if header in normalized_aliases(
                REVISED_FINAL_ALIASES
            ):

                score += 60


            if header in normalized_aliases(
                KOREAN_ALIASES
            ):

                score += 30


            if header in normalized_aliases(
                ORIGINAL_CHINESE_ALIASES
            ):

                score += 20


            if header in normalized_aliases(
                ROW_ALIASES
            ):

                score += 25


        if non_empty >= 2:

            score += min(
                non_empty,
                10,
            )


        if score > best_score:

            best_score = score
            best_row = row_index


    return best_row


# =========================================================
# 检查原翻译文件
# =========================================================

def inspect_original_workbook(
    file_bytes
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )


    result = {}


    for worksheet in workbook.worksheets:

        header_row = detect_header_row(
            worksheet
        )


        korean_candidates = (
            find_column_candidates(
                worksheet=
                    worksheet,

                header_row=
                    header_row,

                column_type=
                    "韩文",
            )
        )


        chinese_candidates = (
            find_column_candidates(
                worksheet=
                    worksheet,

                header_row=
                    header_row,

                column_type=
                    "中文",
            )
        )


        recommended_korean = (
            recommend_candidate(
                korean_candidates
            )
        )


        recommended_chinese = (
            recommend_candidate(
                chinese_candidates
            )
        )


        result[
            worksheet.title
        ] = {

            "工作表":
                worksheet.title,

            "隐藏状态":
                worksheet.sheet_state,

            "行数":
                worksheet.max_row,

            "列数":
                worksheet.max_column,

            "表头行":
                header_row,

            "列名":
                get_headers(
                    worksheet,
                    header_row,
                ),

            "韩文列候选":
                korean_candidates,

            "中文列候选":
                chinese_candidates,

            "推荐韩文列":
                (
                    recommended_korean[
                        "列号"
                    ]
                    if recommended_korean
                    else None
                ),

            "推荐中文列":
                (
                    recommended_chinese[
                        "列号"
                    ]
                    if recommended_chinese
                    else None
                ),

            "ID列候选":
                find_alias_columns(
                    worksheet,
                    header_row,
                    ID_ALIASES,
                ),
        }


    return result


# =========================================================
# 检查部门反馈文件
# =========================================================

def inspect_feedback_workbook(
    file_bytes
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )


    result = {}


    for worksheet in workbook.worksheets:

        header_row = (
            detect_feedback_header_row(
                worksheet
            )
        )


        revised_candidates = (
            find_revised_candidates(
                worksheet,
                header_row,
            )
        )


        recommended_revised = (
            recommend_revised_column(
                revised_candidates
            )
        )


        result[
            worksheet.title
        ] = {

            "工作表":
                worksheet.title,

            "隐藏状态":
                worksheet.sheet_state,

            "行数":
                worksheet.max_row,

            "列数":
                worksheet.max_column,

            "表头行":
                header_row,

            "列名":
                get_headers(
                    worksheet,
                    header_row,
                ),

            "修订后中文列候选":
                revised_candidates,

            "推荐修订后中文列":
                (
                    recommended_revised[
                        "列号"
                    ]
                    if recommended_revised
                    else None
                ),

            "韩文列候选":
                find_alias_columns(
                    worksheet,
                    header_row,
                    KOREAN_ALIASES,
                ),

            "原中文列候选":
                find_alias_columns(
                    worksheet,
                    header_row,
                    ORIGINAL_CHINESE_ALIASES,
                ),

            "ID列候选":
                find_alias_columns(
                    worksheet,
                    header_row,
                    ID_ALIASES,
                ),

            "原表行号列候选":
                find_alias_columns(
                    worksheet,
                    header_row,
                    ROW_ALIASES,
                ),

            "原工作表列候选":
                find_alias_columns(
                    worksheet,
                    header_row,
                    SHEET_ALIASES,
                ),
        }


    return result


# =========================================================
# 选择反馈工作表
# =========================================================

def choose_feedback_sheet(
    workbook,
    sheet_name=None,
):

    if sheet_name:

        if sheet_name not in workbook.sheetnames:

            raise ValueError(
                f"反馈工作表不存在：{sheet_name}"
            )


        return workbook[
            sheet_name
        ]


    # 优先“修订明细”
    if "修订明细" in workbook.sheetnames:

        return workbook[
            "修订明细"
        ]


    candidates = []


    for worksheet in workbook.worksheets:

        header_row = (
            detect_feedback_header_row(
                worksheet
            )
        )


        revised_candidates = (
            find_revised_candidates(
                worksheet,
                header_row,
            )
        )


        recommended = (
            recommend_revised_column(
                revised_candidates
            )
        )


        if recommended:

            candidates.append(
                worksheet.title
            )


    if len(candidates) == 1:

        return workbook[
            candidates[0]
        ]


    raise ValueError(
        (
            "无法唯一确定部门反馈工作表。"
            "请在界面中人工选择反馈工作表。"
        )
    )


# =========================================================
# 查找第一候选列
# =========================================================

def first_alias_column(
    worksheet,
    header_row,
    aliases
):

    candidates = (
        find_alias_columns(
            worksheet,
            header_row,
            aliases,
        )
    )


    if not candidates:
        return None


    return candidates[
        0
    ][
        "列号"
    ]


# =========================================================
# 取得上下文字段列
# =========================================================

def find_context_columns(
    worksheet,
    header_row,
):

    result = {}


    for field_name, aliases in (
        CONTEXT_ALIASES.items()
    ):

        result[
            field_name
        ] = first_alias_column(
            worksheet,
            header_row,
            aliases,
        )


    return result


# =========================================================
# 安全取值
# =========================================================

def get_cell_value(
    worksheet,
    row_index,
    column_index,
):

    if not column_index:
        return ""


    cell = worksheet.cell(
        row=row_index,
        column=column_index,
    )


    if isinstance(
        cell,
        MergedCell,
    ):

        return ""


    return clean_text(
        cell.value
    )


# =========================================================
# 解析原表行号
# =========================================================

def parse_row_number(
    value
):

    text = clean_text(
        value
    )


    if not text:
        return None


    try:

        return int(
            float(
                text
            )
        )


    except Exception:

        return None


# =========================================================
# 构建原文件记录
# =========================================================

def build_original_records(
    worksheet,
    header_row,
    korean_column,
    chinese_column,
    id_column=None,
):

    context_columns = (
        find_context_columns(
            worksheet,
            header_row,
        )
    )


    records = []


    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):

        korean = get_cell_value(
            worksheet,
            row_index,
            korean_column,
        )


        chinese = get_cell_value(
            worksheet,
            row_index,
            chinese_column,
        )


        record_id = get_cell_value(
            worksheet,
            row_index,
            id_column,
        )


        record = {

            "行号":
                row_index,

            "ID":
                record_id,

            "韩文":
                korean,

            "中文":
                chinese,

            "msgctxt":
                get_cell_value(
                    worksheet,
                    row_index,
                    context_columns.get(
                        "msgctxt"
                    ),
                ),

            "references":
                get_cell_value(
                    worksheet,
                    row_index,
                    context_columns.get(
                        "references"
                    ),
                ),

            "Release":
                get_cell_value(
                    worksheet,
                    row_index,
                    context_columns.get(
                        "Release"
                    ),
                ),

            "Content":
                get_cell_value(
                    worksheet,
                    row_index,
                    context_columns.get(
                        "Content"
                    ),
                ),

            "script":
                get_cell_value(
                    worksheet,
                    row_index,
                    context_columns.get(
                        "script"
                    ),
                ),
        }


        records.append(
            record
        )


    return records


# =========================================================
# 建立索引
# =========================================================

def build_indexes(
    records
):

    by_id = {}
    by_korean = {}
    by_row = {}


    for record in records:

        row_index = record[
            "行号"
        ]


        by_row[
            row_index
        ] = record


        record_id = clean_text(
            record.get(
                "ID",
                "",
            )
        )


        if record_id:

            by_id.setdefault(
                record_id,
                []
            ).append(
                record
            )


        korean = clean_text(
            record.get(
                "韩文",
                "",
            )
        )


        if korean:

            by_korean.setdefault(
                korean,
                []
            ).append(
                record
            )


    return {
        "按行号":
            by_row,

        "按ID":
            by_id,

        "按韩文":
            by_korean,
    }


# =========================================================
# 根据反馈上下文缩小候选
# =========================================================

def filter_candidates(
    candidates,
    feedback,
):

    current = list(
        candidates
    )


    # =====================================================
    # 韩文
    # =====================================================

    korean = clean_text(
        feedback.get(
            "韩文",
            "",
        )
    )


    if korean:

        matched = [
            item
            for item in current
            if clean_text(
                item.get(
                    "韩文",
                    "",
                )
            )
            ==
            korean
        ]


        if matched:

            current = matched


    # =====================================================
    # 原中文
    # =====================================================

    original_chinese = clean_text(
        feedback.get(
            "原中文",
            "",
        )
    )


    if original_chinese:

        matched = [
            item
            for item in current
            if clean_text(
                item.get(
                    "中文",
                    "",
                )
            )
            ==
            original_chinese
        ]


        if matched:

            current = matched


    # =====================================================
    # 上下文
    # =====================================================

    for field_name in [
        "msgctxt",
        "references",
        "Release",
        "Content",
        "script",
    ]:

        expected = clean_text(
            feedback.get(
                field_name,
                "",
            )
        )


        if not expected:
            continue


        matched = [
            item
            for item in current
            if clean_text(
                item.get(
                    field_name,
                    "",
                )
            )
            ==
            expected
        ]


        if matched:

            current = matched


    return current


# =========================================================
# 匹配反馈记录
# =========================================================

def match_feedback_record(
    feedback,
    indexes,
    original_sheet_name,
):

    feedback_sheet = clean_text(
        feedback.get(
            "原工作表",
            "",
        )
    )


    # =====================================================
    # 如果报告明确写了其他工作表
    # =====================================================

    if (
        feedback_sheet
        and
        feedback_sheet
        !=
        original_sheet_name
    ):

        return {
            "状态":
                "冲突",

            "匹配方式":
                "",

            "目标":
                None,

            "原因":
                (
                    f"反馈记录指定原工作表“{feedback_sheet}”，"
                    f"当前处理工作表为“{original_sheet_name}”。"
                ),
        }


    # =====================================================
    # 1. 原表行号
    # =====================================================

    row_number = feedback.get(
        "原表行号"
    )


    if row_number:

        candidate = (
            indexes[
                "按行号"
            ].get(
                row_number
            )
        )


        if candidate:

            filtered = (
                filter_candidates(
                    [
                        candidate
                    ],
                    feedback,
                )
            )


            if len(
                filtered
            ) == 1:

                # 如果反馈提供了韩文，
                # 必须严格验证，不能行号对了就盲写。
                feedback_korean = clean_text(
                    feedback.get(
                        "韩文",
                        "",
                    )
                )


                if (
                    feedback_korean
                    and
                    feedback_korean
                    !=
                    clean_text(
                        candidate.get(
                            "韩文",
                            "",
                        )
                    )
                ):

                    return {
                        "状态":
                            "冲突",

                        "匹配方式":
                            "原表行号",

                        "目标":
                            None,

                        "原因":
                            (
                                "原表行号存在，"
                                "但反馈韩文与目标行韩文不一致。"
                            ),
                    }


                return {
                    "状态":
                        "匹配成功",

                    "匹配方式":
                        "原表行号",

                    "目标":
                        candidate,

                    "原因":
                        "",
                }


        return {
            "状态":
                "未匹配",

            "匹配方式":
                "原表行号",

            "目标":
                None,

            "原因":
                (
                    f"原表行号 {row_number} "
                    "无法在当前工作表中可靠匹配。"
                ),
        }


    # =====================================================
    # 2. ID
    # =====================================================

    record_id = clean_text(
        feedback.get(
            "ID",
            "",
        )
    )


    if record_id:

        candidates = (
            indexes[
                "按ID"
            ].get(
                record_id,
                [],
            )
        )


        if candidates:

            filtered = (
                filter_candidates(
                    candidates,
                    feedback,
                )
            )


            if len(
                filtered
            ) == 1:

                return {
                    "状态":
                        "匹配成功",

                    "匹配方式":
                        (
                            "唯一ID"
                            if len(
                                candidates
                            ) == 1
                            else
                            "ID+上下文"
                        ),

                    "目标":
                        filtered[
                            0
                        ],

                    "原因":
                        "",
                }


            if len(
                filtered
            ) > 1:

                return {
                    "状态":
                        "冲突",

                    "匹配方式":
                        "ID+上下文",

                    "目标":
                        None,

                    "原因":
                        (
                            "相同ID仍匹配到多条记录，"
                            "上下文不足，禁止自动写回。"
                        ),
                }


    # =====================================================
    # 3. 韩文
    # =====================================================

    korean = clean_text(
        feedback.get(
            "韩文",
            "",
        )
    )


    if korean:

        candidates = (
            indexes[
                "按韩文"
            ].get(
                korean,
                [],
            )
        )


        if not candidates:

            return {
                "状态":
                    "未匹配",

                "匹配方式":
                    "韩文",

                "目标":
                    None,

                "原因":
                    "反馈韩文在原文件中不存在。",
            }


        filtered = (
            filter_candidates(
                candidates,
                feedback,
            )
        )


        if len(
            filtered
        ) == 1:

            if len(
                candidates
            ) == 1:

                method = (
                    "唯一韩文"
                )

            else:

                method = (
                    "韩文+原中文/上下文"
                )


            return {
                "状态":
                    "匹配成功",

                "匹配方式":
                    method,

                "目标":
                    filtered[
                        0
                    ],

                "原因":
                    "",
            }


        if len(
            filtered
        ) > 1:

            return {
                "状态":
                    "冲突",

                "匹配方式":
                    "重复韩文",

                "目标":
                    None,

                "原因":
                    (
                        "同一韩文在原文件中出现多次，"
                        "现有原中文/上下文仍无法唯一确定。"
                        "系统禁止把一条部门修订扩散到所有重复韩文。"
                    ),
            }


    # =====================================================
    # 无任何可靠关键字段
    # =====================================================

    return {
        "状态":
            "未匹配",

        "匹配方式":
            "",

        "目标":
            None,

        "原因":
            (
                "反馈记录缺少可用于精确匹配的"
                "原表行号、ID或韩文。"
            ),
    }


# =========================================================
# 删除旧报告
# =========================================================

def remove_generated_sheets(
    workbook
):

    for sheet_name in GENERATED_SHEETS:

        if sheet_name in workbook.sheetnames:

            workbook.remove(
                workbook[
                    sheet_name
                ]
            )


# =========================================================
# 生成统计
# =========================================================

def new_stats():

    return {

        "反馈总行数":
            0,

        "有效修订行":
            0,

        "成功写回":
            0,

        "无需修改":
            0,

        "空修订跳过":
            0,

        "未匹配":
            0,

        "冲突待确认":
            0,

        "格式异常":
            0,

        "结构保护":
            0,

        "原表行号匹配":
            0,

        "唯一ID匹配":
            0,

        "ID+上下文匹配":
            0,

        "唯一韩文匹配":
            0,

        "韩文+原中文/上下文匹配":
            0,

        "重复韩文未扩散":
            0,
    }


# =========================================================
# 添加报告页
# =========================================================

def add_reports(
    workbook,
    stats,
    details,
    config,
):

    remove_generated_sheets(
        workbook
    )


    # =====================================================
    # 摘要
    # =====================================================

    summary = workbook.create_sheet(
        "部门反馈修订摘要"
    )


    summary.append(
        [
            "项目",
            "结果",
        ]
    )


    summary.append(
        [
            "原文件工作表",
            config.get(
                "原工作表",
                "",
            ),
        ]
    )


    summary.append(
        [
            "部门反馈工作表",
            config.get(
                "反馈工作表",
                "",
            ),
        ]
    )


    summary.append(
        [
            "韩文列",
            config.get(
                "原韩文列",
                "",
            ),
        ]
    )


    summary.append(
        [
            "原中文列",
            config.get(
                "原中文列",
                "",
            ),
        ]
    )


    summary.append(
        [
            "部门修订字段",
            config.get(
                "反馈修订列",
                "",
            ),
        ]
    )


    for key, value in stats.items():

        summary.append(
            [
                key,
                value,
            ]
        )


    accounted = (
        stats[
            "成功写回"
        ]
        +
        stats[
            "无需修改"
        ]
        +
        stats[
            "未匹配"
        ]
        +
        stats[
            "冲突待确认"
        ]
        +
        stats[
            "格式异常"
        ]
        +
        stats[
            "结构保护"
        ]
    )


    summary.append(
        [
            "有效修订核算数",
            accounted,
        ]
    )


    summary.append(
        [
            "核算是否一致",
            (
                "通过"
                if accounted
                ==
                stats[
                    "有效修订行"
                ]
                else
                "不通过"
            ),
        ]
    )


    format_report_sheet(
        summary
    )


    # =====================================================
    # 全部明细
    # =====================================================

    detail_sheet = workbook.create_sheet(
        "部门反馈修订明细"
    )


    write_dict_rows(
        detail_sheet,
        details,
    )


    format_report_sheet(
        detail_sheet
    )


    # =====================================================
    # 未匹配
    # =====================================================

    unmatched = [
        item
        for item in details
        if item.get(
            "处理结果"
        )
        ==
        "未匹配"
    ]


    unmatched_sheet = workbook.create_sheet(
        "未匹配反馈"
    )


    write_dict_rows(
        unmatched_sheet,
        unmatched,
    )


    format_report_sheet(
        unmatched_sheet
    )


    # =====================================================
    # 冲突
    # =====================================================

    conflicts = [
        item
        for item in details
        if item.get(
            "处理结果"
        )
        in {
            "冲突待确认",
            "结构保护",
        }
    ]


    conflict_sheet = workbook.create_sheet(
        "冲突待确认"
    )


    write_dict_rows(
        conflict_sheet,
        conflicts,
    )


    format_report_sheet(
        conflict_sheet
    )


    # =====================================================
    # 格式异常
    # =====================================================

    format_errors = [
        item
        for item in details
        if item.get(
            "处理结果"
        )
        ==
        "格式异常"
    ]


    format_sheet = workbook.create_sheet(
        "格式异常_反馈修订"
    )


    write_dict_rows(
        format_sheet,
        format_errors,
    )


    format_report_sheet(
        format_sheet
    )


# =========================================================
# 部门反馈修订主入口
# =========================================================

def process_department_feedback_revision(
    original_file_bytes,
    feedback_file_bytes,
    original_sheet_name=None,
    feedback_sheet_name=None,
    original_korean_column=None,
    original_chinese_column=None,
    feedback_revised_column=None,
    original_id_column=None,
    feedback_id_column=None,
    feedback_korean_column=None,
    feedback_original_chinese_column=None,
    feedback_row_column=None,
    feedback_source_sheet_column=None,
):
    """
    返回：

    output_bytes
    stats
    details
    config

    注意：

    本模块不调用AI。

    它只执行部门已经确认的修订。
    """

    # =====================================================
    # 读取两个Excel
    # =====================================================

    original_workbook = load_workbook(
        BytesIO(
            original_file_bytes
        ),
        data_only=False,
    )


    feedback_workbook = load_workbook(
        BytesIO(
            feedback_file_bytes
        ),
        data_only=False,
    )


    # =====================================================
    # 原文件工作表
    # =====================================================

    original_sheet = choose_sheet(
        workbook=
            original_workbook,

        sheet_name=
            original_sheet_name,
    )


    original_header_row = (
        detect_header_row(
            original_sheet
        )
    )


    # =====================================================
    # 原韩文列
    # =====================================================

    original_korean_column = (
        choose_korean_column(
            worksheet=
                original_sheet,

            header_row=
                original_header_row,

            korean_column=
                original_korean_column,
        )
    )


    # =====================================================
    # 原中文列
    # =====================================================

    if original_chinese_column is not None:

        original_chinese_column = (
            resolve_column(
                worksheet=
                    original_sheet,

                header_row=
                    original_header_row,

                column_value=
                    original_chinese_column,
            )
        )

    else:

        chinese_candidates = (
            find_column_candidates(
                worksheet=
                    original_sheet,

                header_row=
                    original_header_row,

                column_type=
                    "中文",
            )
        )


        recommended_chinese = (
            recommend_candidate(
                chinese_candidates
            )
        )


        if not recommended_chinese:

            raise ValueError(
                (
                    "无法可靠确定原文件中文列。"
                    "请在界面中人工选择需要修订的中文列。"
                )
            )


        original_chinese_column = (
            recommended_chinese[
                "列号"
            ]
        )


    if (
        original_korean_column
        ==
        original_chinese_column
    ):

        raise ValueError(
            "原文件韩文列与中文列不能相同。"
        )


    # =====================================================
    # 原ID列
    # =====================================================

    if original_id_column is not None:

        original_id_column = (
            resolve_column(
                worksheet=
                    original_sheet,

                header_row=
                    original_header_row,

                column_value=
                    original_id_column,
            )
        )

    else:

        original_id_column = (
            first_alias_column(
                original_sheet,
                original_header_row,
                ID_ALIASES,
            )
        )


    # =====================================================
    # 反馈工作表
    # =====================================================

    feedback_sheet = (
        choose_feedback_sheet(
            workbook=
                feedback_workbook,

            sheet_name=
                feedback_sheet_name,
        )
    )


    feedback_header_row = (
        detect_feedback_header_row(
            feedback_sheet
        )
    )


    # =====================================================
    # 修订后中文列
    # =====================================================

    if feedback_revised_column is not None:

        feedback_revised_column = (
            resolve_column(
                worksheet=
                    feedback_sheet,

                header_row=
                    feedback_header_row,

                column_value=
                    feedback_revised_column,
            )
        )

    else:

        revised_candidates = (
            find_revised_candidates(
                feedback_sheet,
                feedback_header_row,
            )
        )


        recommended_revised = (
            recommend_revised_column(
                revised_candidates
            )
        )


        if not recommended_revised:

            raise ValueError(
                (
                    "没有找到唯一可靠的部门“修订后中文”字段。"
                    "请人工选择部门最终修订列。"
                )
            )


        feedback_revised_column = (
            recommended_revised[
                "列号"
            ]
        )


    # =====================================================
    # 阻止把“建议中文”误当部门最终修订
    # =====================================================

    revised_header = clean_text(
        feedback_sheet.cell(
            row=feedback_header_row,
            column=feedback_revised_column,
        ).value
    )


    if normalize_header(
        revised_header
    ) in normalized_aliases(
        SUGGESTION_ALIASES
    ):

        raise ValueError(
            (
                f"当前选择的“{revised_header}”属于建议字段，"
                "不是部门最终确认字段。"
                "部门反馈修订模式默认禁止把AI建议直接写回原文件。"
            )
        )


    # =====================================================
    # 自动识别反馈字段
    # =====================================================

    def resolve_optional_feedback_column(
        provided,
        aliases,
    ):

        if provided is not None:

            return resolve_column(
                worksheet=
                    feedback_sheet,

                header_row=
                    feedback_header_row,

                column_value=
                    provided,
            )


        return first_alias_column(
            feedback_sheet,
            feedback_header_row,
            aliases,
        )


    feedback_id_column = (
        resolve_optional_feedback_column(
            feedback_id_column,
            ID_ALIASES,
        )
    )


    feedback_korean_column = (
        resolve_optional_feedback_column(
            feedback_korean_column,
            KOREAN_ALIASES,
        )
    )


    feedback_original_chinese_column = (
        resolve_optional_feedback_column(
            feedback_original_chinese_column,
            ORIGINAL_CHINESE_ALIASES,
        )
    )


    feedback_row_column = (
        resolve_optional_feedback_column(
            feedback_row_column,
            ROW_ALIASES,
        )
    )


    feedback_source_sheet_column = (
        resolve_optional_feedback_column(
            feedback_source_sheet_column,
            SHEET_ALIASES,
        )
    )


    feedback_context_columns = (
        find_context_columns(
            feedback_sheet,
            feedback_header_row,
        )
    )


    # =====================================================
    # 构建原记录和索引
    # =====================================================

    original_records = (
        build_original_records(
            worksheet=
                original_sheet,

            header_row=
                original_header_row,

            korean_column=
                original_korean_column,

            chinese_column=
                original_chinese_column,

            id_column=
                original_id_column,
        )
    )


    indexes = (
        build_indexes(
            original_records
        )
    )


    # =====================================================
    # 配置
    # =====================================================

    config = {

        "原工作表":
            original_sheet.title,

        "反馈工作表":
            feedback_sheet.title,

        "原韩文列":
            clean_text(
                original_sheet.cell(
                    row=original_header_row,
                    column=original_korean_column,
                ).value
            ),

        "原中文列":
            clean_text(
                original_sheet.cell(
                    row=original_header_row,
                    column=original_chinese_column,
                ).value
            ),

        "反馈修订列":
            revised_header,
    }


    stats = new_stats()

    details = []


    # =====================================================
    # 逐行处理部门反馈
    # =====================================================

    for feedback_row in range(
        feedback_header_row + 1,
        feedback_sheet.max_row + 1,
    ):

        stats[
            "反馈总行数"
        ] += 1


        revised_chinese = (
            get_cell_value(
                feedback_sheet,
                feedback_row,
                feedback_revised_column,
            )
        )


        # =================================================
        # 修订字段为空
        # =================================================

        if not revised_chinese:

            stats[
                "空修订跳过"
            ] += 1

            continue


        stats[
            "有效修订行"
        ] += 1


        feedback_data = {

            "反馈行号":
                feedback_row,

            "原表行号":
                parse_row_number(
                    get_cell_value(
                        feedback_sheet,
                        feedback_row,
                        feedback_row_column,
                    )
                ),

            "原工作表":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_source_sheet_column,
                ),

            "ID":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_id_column,
                ),

            "韩文":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_korean_column,
                ),

            "原中文":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_original_chinese_column,
                ),

            "修订后中文":
                revised_chinese,

            "msgctxt":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_context_columns.get(
                        "msgctxt"
                    ),
                ),

            "references":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_context_columns.get(
                        "references"
                    ),
                ),

            "Release":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_context_columns.get(
                        "Release"
                    ),
                ),

            "Content":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_context_columns.get(
                        "Content"
                    ),
                ),

            "script":
                get_cell_value(
                    feedback_sheet,
                    feedback_row,
                    feedback_context_columns.get(
                        "script"
                    ),
                ),
        }


        # =================================================
        # 精确匹配
        # =================================================

        match_result = (
            match_feedback_record(
                feedback=
                    feedback_data,

                indexes=
                    indexes,

                original_sheet_name=
                    original_sheet.title,
            )
        )


        match_status = (
            match_result.get(
                "状态",
                "",
            )
        )


        match_method = (
            match_result.get(
                "匹配方式",
                "",
            )
        )


        # =================================================
        # 未匹配
        # =================================================

        if match_status == "未匹配":

            stats[
                "未匹配"
            ] += 1


            details.append(
                {
                    **feedback_data,

                    "目标原表行号":
                        "",

                    "匹配方式":
                        match_method,

                    "写回前中文":
                        "",

                    "最终写回中文":
                        "",

                    "格式检查":
                        "未执行",

                    "处理结果":
                        "未匹配",

                    "原因":
                        match_result.get(
                            "原因",
                            "",
                        ),
                }
            )


            continue


        # =================================================
        # 冲突
        # =================================================

        if match_status == "冲突":

            stats[
                "冲突待确认"
            ] += 1


            if match_method == "重复韩文":

                stats[
                    "重复韩文未扩散"
                ] += 1


            details.append(
                {
                    **feedback_data,

                    "目标原表行号":
                        "",

                    "匹配方式":
                        match_method,

                    "写回前中文":
                        "",

                    "最终写回中文":
                        "",

                    "格式检查":
                        "未执行",

                    "处理结果":
                        "冲突待确认",

                    "原因":
                        match_result.get(
                            "原因",
                            "",
                        ),
                }
            )


            continue


        # =================================================
        # 成功找到唯一目标
        # =================================================

        target_record = (
            match_result[
                "目标"
            ]
        )


        target_row = (
            target_record[
                "行号"
            ]
        )


        korean_text = clean_text(
            target_record.get(
                "韩文",
                "",
            )
        )


        before_chinese = get_cell_value(
            original_sheet,
            target_row,
            original_chinese_column,
        )


        # =================================================
        # 匹配方式统计
        # =================================================

        if match_method == "原表行号":

            stats[
                "原表行号匹配"
            ] += 1


        elif match_method == "唯一ID":

            stats[
                "唯一ID匹配"
            ] += 1


        elif match_method == "ID+上下文":

            stats[
                "ID+上下文匹配"
            ] += 1


        elif match_method == "唯一韩文":

            stats[
                "唯一韩文匹配"
            ] += 1


        elif match_method == "韩文+原中文/上下文":

            stats[
                "韩文+原中文/上下文匹配"
            ] += 1


        # =================================================
        # 目标单元格结构保护
        # =================================================

        target_cell = original_sheet.cell(
            row=target_row,
            column=original_chinese_column,
        )


        if (
            isinstance(
                target_cell,
                MergedCell,
            )
            or
            is_formula_cell(
                target_cell
            )
        ):

            stats[
                "结构保护"
            ] += 1


            details.append(
                {
                    **feedback_data,

                    "目标原表行号":
                        target_row,

                    "匹配方式":
                        match_method,

                    "写回前中文":
                        before_chinese,

                    "最终写回中文":
                        "",

                    "格式检查":
                        "未执行",

                    "处理结果":
                        "结构保护",

                    "原因":
                        (
                            "目标中文单元格为公式或合并单元格，"
                            "系统禁止自动写回。"
                        ),
                }
            )


            continue


        # =================================================
        # 格式QA
        # =================================================

        qa_result = check_format(
            korean_text,
            revised_chinese,
        )


        qa_summary = format_qa_summary(
            korean_text,
            revised_chinese,
        )


        # =================================================
        # 部门修订存在格式风险
        #
        # 不自动写回
        # =================================================

        if not qa_result.get(
            "通过",
            False,
        ):

            stats[
                "格式异常"
            ] += 1


            details.append(
                {
                    **feedback_data,

                    "目标原表行号":
                        target_row,

                    "匹配方式":
                        match_method,

                    "写回前中文":
                        before_chinese,

                    "最终写回中文":
                        "",

                    "格式检查":
                        qa_summary,

                    "处理结果":
                        "格式异常",

                    "原因":
                        (
                            "部门修订后中文未通过程序级格式QA，"
                            "为避免破坏数字、占位符、标签等内容，"
                            "系统未自动写回。"
                        ),
                }
            )


            continue


        # =================================================
        # 原中文已经相同
        # =================================================

        if (
            clean_text(
                before_chinese
            )
            ==
            clean_text(
                revised_chinese
            )
        ):

            stats[
                "无需修改"
            ] += 1


            details.append(
                {
                    **feedback_data,

                    "目标原表行号":
                        target_row,

                    "匹配方式":
                        match_method,

                    "写回前中文":
                        before_chinese,

                    "最终写回中文":
                        before_chinese,

                    "格式检查":
                        qa_summary,

                    "处理结果":
                        "无需修改",

                    "原因":
                        (
                            "原文件中文已经与部门修订后中文一致。"
                        ),
                }
            )


            continue


        # =================================================
        # 正式写回
        # =================================================

        target_cell.value = (
            revised_chinese
        )


        # 同步内存记录，
        # 防止后续反馈继续使用旧中文匹配。
        target_record[
            "中文"
        ] = revised_chinese


        stats[
            "成功写回"
        ] += 1


        details.append(
            {
                **feedback_data,

                "目标原表行号":
                    target_row,

                "匹配方式":
                    match_method,

                "写回前中文":
                    before_chinese,

                "最终写回中文":
                    revised_chinese,

                "格式检查":
                    qa_summary,

                "处理结果":
                    "成功写回",

                "原因":
                    (
                        "已按照部门确认的修订后中文"
                        "精确写回对应原文件记录。"
                    ),
            }
        )


    # =====================================================
    # 核算校验
    # =====================================================

    accounted = (
        stats[
            "成功写回"
        ]
        +
        stats[
            "无需修改"
        ]
        +
        stats[
            "未匹配"
        ]
        +
        stats[
            "冲突待确认"
        ]
        +
        stats[
            "格式异常"
        ]
        +
        stats[
            "结构保护"
        ]
    )


    if accounted != stats[
        "有效修订行"
    ]:

        raise RuntimeError(
            (
                "部门反馈修订核算失败："
                f"有效修订行={stats['有效修订行']}，"
                f"已核算={accounted}。"
                "程序已停止输出，避免生成不完整修订文件。"
            )
        )


    # =====================================================
    # 添加审计报告
    # =====================================================

    add_reports(
        workbook=
            original_workbook,

        stats=
            stats,

        details=
            details,

        config=
            config,
    )


    # =====================================================
    # 最后保存
    # =====================================================

    output = BytesIO()


    original_workbook.save(
        output
    )


    output.seek(
        0
    )


    return (
        output.getvalue(),
        stats,
        details,
        config,
    )


# =========================================================
# 独立测试
# =========================================================

if __name__ == "__main__":

    print(
        "部门反馈修订模块加载成功。"
    )