from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from strict_reviewer import (
    review_text,
)

from document_processor import (
    BatchAIUnavailableError,
    contains_korean,
)

from xlsx_translator import (
    ColumnSelectionRequired,
    detect_header_row,
    get_headers,
    find_column_candidates,
    recommend_candidate,
    resolve_column,
    choose_sheet,
    choose_korean_column,
    build_header_map,
    build_excel_context,
    cell_text,
    is_formula_cell,
    format_report_sheet,
    write_dict_rows,
)


# =========================================================
# XLSX 批量严格审校
#
# 用途：
#
# 对已经有：
#
# 韩文原文 + 现有中文译文
#
# 的Excel执行逐行严格审校。
#
# 固定优先级：
#
# 角色名库
# ＞ UWO正式主库
# ＞ Quest正式主库
# ＞ 当前Excel上下文
# ＞ 游戏内历史参考
# ＞ 保守判断
#
# 核心原则：
#
# 1. 原业务工作表不修改
# 2. 现有中文不覆盖
# 3. 唯一正式完整句由程序直接判断
# 4. 能程序判断时不调用AI
# 5. 无法可靠判断时才调用AI
# 6. AI不得自由润色
# 7. 格式QA由程序最终检查
# 8. 多译法 / 冲突 → 人工确认
# 9. 第一次Model API致命错误立即终止整批
# 10. 致命错误时不生成半成品Excel
#
# 成功输出：
#
# 原工作簿
# +
# 严格审校摘要
# 严格审校明细
# 建议修改
# 需人工确认_审校
# 格式异常_审校
# =========================================================


GENERATED_REVIEW_SHEETS = [
    "严格审校摘要",
    "严格审校明细",
    "建议修改",
    "需人工确认_审校",
    "格式异常_审校",
]


ID_HEADER_ALIASES = [
    "id",
    "stringid",
    "string_id",
    "key",
    "文本id",
    "字符串id",
    "msgctxt",
]


# =========================================================
# 表头标准化
# =========================================================

def normalize_header(value):

    if value is None:
        return ""

    return str(
        value
    ).strip().lower()


# =========================================================
# 取得记录ID
# =========================================================

def get_record_id(
    worksheet,
    row_index,
    header_map,
):

    for alias in ID_HEADER_ALIASES:

        column_index = header_map.get(
            normalize_header(
                alias
            )
        )

        if not column_index:
            continue

        cell = worksheet.cell(
            row=row_index,
            column=column_index,
        )

        if isinstance(
            cell,
            MergedCell,
        ):
            continue

        value = cell.value

        if value is None:
            continue

        text = str(
            value
        ).strip()

        if text:
            return text

    return ""


# =========================================================
# XLSX严格审校结构检查
# =========================================================

def inspect_xlsx_for_review(
    file_bytes
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )

    result = {}

    for worksheet in workbook.worksheets:

        header_row = detect_header_row(
            worksheet
        )

        korean_candidates = (
            find_column_candidates(
                worksheet=worksheet,
                header_row=header_row,
                column_type="韩文",
            )
        )

        chinese_candidates = (
            find_column_candidates(
                worksheet=worksheet,
                header_row=header_row,
                column_type="中文",
            )
        )

        recommended_korean = (
            recommend_candidate(
                korean_candidates
            )
        )

        recommended_chinese = (
            recommend_candidate(
                chinese_candidates
            )
        )

        result[
            worksheet.title
        ] = {

            "工作表":
                worksheet.title,

            "隐藏状态":
                worksheet.sheet_state,

            "行数":
                worksheet.max_row,

            "列数":
                worksheet.max_column,

            "表头行":
                header_row,

            "列名":
                get_headers(
                    worksheet,
                    header_row,
                ),

            "韩文列候选":
                korean_candidates,

            "中文列候选":
                chinese_candidates,

            "推荐韩文列":
                (
                    recommended_korean[
                        "列号"
                    ]
                    if recommended_korean
                    else None
                ),

            "推荐中文列":
                (
                    recommended_chinese[
                        "列号"
                    ]
                    if recommended_chinese
                    else None
                ),
        }

    return result


# =========================================================
# 选择现有中文列
# =========================================================

def choose_review_chinese_column(
    worksheet,
    header_row,
    source_column,
    chinese_column=None,
):

    # 用户明确指定
    if chinese_column is not None:

        resolved = resolve_column(
            worksheet=worksheet,
            header_row=header_row,
            column_value=chinese_column,
        )

        if resolved == source_column:

            raise ValueError(
                "现有中文列不能与韩文原文列相同。"
            )

        return resolved

    # 自动检测
    candidates = (
        find_column_candidates(
            worksheet=worksheet,
            header_row=header_row,
            column_type="中文",
        )
    )

    candidates = [
        item
        for item in candidates
        if item.get(
            "列号"
        )
        != source_column
    ]

    recommended = (
        recommend_candidate(
            candidates
        )
    )

    if recommended:

        return recommended[
            "列号"
        ]

    raise ColumnSelectionRequired(
        (
            "严格审校需要明确的现有中文列，"
            "当前无法可靠自动判断。"
            "请人工选择需要审校的中文译文列。"
        ),
        column_type="中文",
        candidates=candidates,
    )


# =========================================================
# 统计
# =========================================================

def new_review_stats():

    return {

        "总数据行":
            0,

        "韩文有效行":
            0,

        "现有中文为空":
            0,

        "通过":
            0,

        "建议修改":
            0,

        "人工确认":
            0,

        "程序精确审校":
            0,

        "AI严格审校":
            0,

        "AI审校失败":
            0,

        "角色名命中":
            0,

        "格式异常":
            0,

        "精确匹配":
            0,

        "相似匹配":
            0,

        "候选匹配":
            0,

        "未确认匹配":
            0,

        "非韩文跳过":
            0,

        "空原文跳过":
            0,

        "公式人工确认":
            0,

        "合并单元格人工确认":
            0,
    }


# =========================================================
# 更新统计
# =========================================================

def update_review_stats(
    stats,
    result,
):

    verdict = str(
        result.get(
            "审校结论",
            "",
        )
    )

    if verdict == "通过":

        stats[
            "通过"
        ] += 1

    elif verdict == "建议修改":

        stats[
            "建议修改"
        ] += 1

    elif verdict == "人工确认":

        stats[
            "人工确认"
        ] += 1


    # =====================================================
    # 审校方式
    # =====================================================

    method = str(
        result.get(
            "审校方式",
            "",
        )
    )

    if method == "程序精确审校":

        stats[
            "程序精确审校"
        ] += 1

    elif method == "知识库约束下AI严格审校":

        stats[
            "AI严格审校"
        ] += 1

    elif method == "AI严格审校失败":

        stats[
            "AI审校失败"
        ] += 1


    # =====================================================
    # 角色名
    # =====================================================

    if str(
        result.get(
            "角色名命中",
            "",
        )
    ).strip():

        stats[
            "角色名命中"
        ] += 1


    # =====================================================
    # 匹配状态
    # =====================================================

    match_status = str(
        result.get(
            "匹配状态",
            "",
        )
    )

    if match_status == "精确":

        stats[
            "精确匹配"
        ] += 1

    elif match_status == "相似":

        stats[
            "相似匹配"
        ] += 1

    elif match_status == "候选":

        stats[
            "候选匹配"
        ] += 1

    elif match_status == "未确认":

        stats[
            "未确认匹配"
        ] += 1


    # =====================================================
    # 格式异常
    # =====================================================

    existing_qa = str(
        result.get(
            "现有译文格式检查",
            "",
        )
    )

    suggested_qa = str(
        result.get(
            "建议译文格式检查",
            "",
        )
    )

    if (
        existing_qa.startswith(
            "不通过"
        )
        or
        suggested_qa.startswith(
            "不通过"
        )
    ):

        stats[
            "格式异常"
        ] += 1


# =========================================================
# API致命错误
# =========================================================

def raise_if_fatal_review_error(
    result,
    location,
):

    if not result.get(
        "致命错误",
        False,
    ):
        return

    error_type = str(
        result.get(
            "错误类型",
            "AIReviewError",
        )
        or
        "AIReviewError"
    )

    error_message = str(
        result.get(
            "错误",
            "",
        )
        or
        result.get(
            "人工确认原因",
            "",
        )
        or
        "AI模型严格审校当前不可用。"
    ).strip()

    raise BatchAIUnavailableError(
        message=(
            f"严格审校在“{location}”发生致命AI错误。"
            f"错误类型：{error_type}。"
            f"{error_message} "
            "系统已经停止后续所有审校请求，"
            "没有生成部分完成的严格审校文件。"
        ),
        location=location,
        error_type=error_type,
    )


# =========================================================
# 删除旧严格审校报告
# =========================================================

def remove_old_review_sheets(
    workbook
):

    for sheet_name in (
        GENERATED_REVIEW_SHEETS
    ):

        if sheet_name in workbook.sheetnames:

            workbook.remove(
                workbook[
                    sheet_name
                ]
            )


# =========================================================
# 添加严格审校报告
# =========================================================

def add_review_reports(
    workbook,
    stats,
    details,
    config,
):

    remove_old_review_sheets(
        workbook
    )


    # =====================================================
    # 摘要
    # =====================================================

    summary_sheet = workbook.create_sheet(
        "严格审校摘要"
    )

    summary_sheet.append(
        [
            "项目",
            "结果",
        ]
    )

    summary_sheet.append(
        [
            "处理工作表",
            config.get(
                "工作表",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "表头行",
            config.get(
                "表头行",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "韩文原文列",
            config.get(
                "韩文列",
                "",
            ),
        ]
    )

    summary_sheet.append(
        [
            "现有中文列",
            config.get(
                "中文列",
                "",
            ),
        ]
    )

    for key, value in stats.items():

        summary_sheet.append(
            [
                key,
                value,
            ]
        )


    reviewed_total = (
        stats[
            "通过"
        ]
        +
        stats[
            "建议修改"
        ]
        +
        stats[
            "人工确认"
        ]
    )

    if reviewed_total > 0:

        pass_rate = (
            stats[
                "通过"
            ]
            /
            reviewed_total
            *
            100
        )

        modify_rate = (
            stats[
                "建议修改"
            ]
            /
            reviewed_total
            *
            100
        )

        manual_rate = (
            stats[
                "人工确认"
            ]
            /
            reviewed_total
            *
            100
        )

    else:

        pass_rate = 0
        modify_rate = 0
        manual_rate = 0


    summary_sheet.append(
        [
            "审校通过率",
            f"{pass_rate:.2f}%",
        ]
    )

    summary_sheet.append(
        [
            "建议修改率",
            f"{modify_rate:.2f}%",
        ]
    )

    summary_sheet.append(
        [
            "人工确认率",
            f"{manual_rate:.2f}%",
        ]
    )

    format_report_sheet(
        summary_sheet
    )


    # =====================================================
    # 完整审校明细
    # =====================================================

    detail_sheet = workbook.create_sheet(
        "严格审校明细"
    )

    write_dict_rows(
        detail_sheet,
        details,
    )

    format_report_sheet(
        detail_sheet
    )


    # =====================================================
    # 建议修改
    # =====================================================

    modify_rows = [
        row
        for row in details
        if row.get(
            "审校结论"
        )
        ==
        "建议修改"
    ]

    modify_sheet = workbook.create_sheet(
        "建议修改"
    )

    write_dict_rows(
        modify_sheet,
        modify_rows,
    )

    format_report_sheet(
        modify_sheet
    )


    # =====================================================
    # 人工确认
    # =====================================================

    manual_rows = [
        row
        for row in details
        if (
            row.get(
                "审校结论"
            )
            ==
            "人工确认"

            or

            row.get(
                "需人工确认",
                False,
            )
        )
    ]

    manual_sheet = workbook.create_sheet(
        "需人工确认_审校"
    )

    write_dict_rows(
        manual_sheet,
        manual_rows,
    )

    format_report_sheet(
        manual_sheet
    )


    # =====================================================
    # 格式异常
    # =====================================================

    format_rows = []

    for row in details:

        existing_qa = str(
            row.get(
                "现有译文格式检查",
                "",
            )
        )

        suggested_qa = str(
            row.get(
                "建议译文格式检查",
                "",
            )
        )

        if (
            existing_qa.startswith(
                "不通过"
            )
            or
            suggested_qa.startswith(
                "不通过"
            )
        ):

            format_rows.append(
                row
            )


    format_sheet = workbook.create_sheet(
        "格式异常_审校"
    )

    write_dict_rows(
        format_sheet,
        format_rows,
    )

    format_report_sheet(
        format_sheet
    )


# =========================================================
# 特殊人工确认记录
# =========================================================

def make_special_manual_result(
    worksheet,
    row_index,
    record_id,
    korean_text,
    existing_chinese,
    reason,
):

    return {

        "工作表":
            worksheet.title,

        "原表行号":
            row_index,

        "ID":
            record_id,

        "韩文原文":
            korean_text,

        "现有中文":
            existing_chinese,

        "建议中文":
            existing_chinese,

        "审校结论":
            "人工确认",

        "角色名命中":
            "",

        "术语命中":
            "",

        "参考知识库":
            "",

        "匹配状态":
            "未确认",

        "问题说明":
            reason,

        "修改原因":
            "",

        "现有译文格式检查":
            "未执行",

        "建议译文格式检查":
            "未执行",

        "需人工确认":
            True,

        "人工确认原因":
            reason,

        "审校方式":
            "程序安全保护",

        "AI成功":
            False,

        "错误类型":
            "",

        "错误":
            "",

        "致命错误":
            False,
    }


# =========================================================
# XLSX批量严格审校主入口
# =========================================================

def process_xlsx_strict_review(
    file_bytes,
    knowledge_base,
    sheet_name=None,
    korean_column=None,
    chinese_column=None,
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )


    # =====================================================
    # 工作表
    # =====================================================

    worksheet = choose_sheet(
        workbook=workbook,
        sheet_name=sheet_name,
    )

    header_row = detect_header_row(
        worksheet
    )


    # =====================================================
    # 韩文列
    # =====================================================

    source_column = choose_korean_column(
        worksheet=worksheet,
        header_row=header_row,
        korean_column=korean_column,
    )


    # =====================================================
    # 中文列
    # =====================================================

    target_column = (
        choose_review_chinese_column(
            worksheet=worksheet,
            header_row=header_row,
            source_column=source_column,
            chinese_column=chinese_column,
        )
    )

    if source_column == target_column:

        raise ValueError(
            "韩文原文列和现有中文列不能相同。"
        )


    # =====================================================
    # 配置
    # =====================================================

    source_header = worksheet.cell(
        row=header_row,
        column=source_column,
    ).value

    target_header = worksheet.cell(
        row=header_row,
        column=target_column,
    ).value


    config = {

        "工作表":
            worksheet.title,

        "表头行":
            header_row,

        "韩文列号":
            source_column,

        "韩文列":
            (
                str(
                    source_header
                )
                if source_header is not None
                else
                get_column_letter(
                    source_column
                )
            ),

        "中文列号":
            target_column,

        "中文列":
            (
                str(
                    target_header
                )
                if target_header is not None
                else
                get_column_letter(
                    target_column
                )
            ),
    }


    header_map = build_header_map(
        worksheet=worksheet,
        header_row=header_row,
    )

    stats = new_review_stats()

    details = []


    # =====================================================
    # 逐行严格审校
    # =====================================================

    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):

        stats[
            "总数据行"
        ] += 1


        source_cell = worksheet.cell(
            row=row_index,
            column=source_column,
        )

        target_cell = worksheet.cell(
            row=row_index,
            column=target_column,
        )


        record_id = get_record_id(
            worksheet=worksheet,
            row_index=row_index,
            header_map=header_map,
        )


        # =================================================
        # 合并单元格
        # =================================================

        if (
            isinstance(
                source_cell,
                MergedCell,
            )
            or
            isinstance(
                target_cell,
                MergedCell,
            )
        ):

            stats[
                "合并单元格人工确认"
            ] += 1

            manual_result = (
                make_special_manual_result(
                    worksheet=worksheet,
                    row_index=row_index,
                    record_id=record_id,
                    korean_text=(
                        cell_text(
                            source_cell
                        )
                        if not isinstance(
                            source_cell,
                            MergedCell,
                        )
                        else ""
                    ),
                    existing_chinese=(
                        cell_text(
                            target_cell
                        )
                        if not isinstance(
                            target_cell,
                            MergedCell,
                        )
                        else ""
                    ),
                    reason=(
                        "当前行涉及合并单元格，"
                        "为保护Excel结构，"
                        "严格审校标记为人工确认。"
                    ),
                )
            )

            details.append(
                manual_result
            )

            stats[
                "人工确认"
            ] += 1

            continue


        # =================================================
        # 公式
        # =================================================

        if (
            is_formula_cell(
                source_cell
            )
            or
            is_formula_cell(
                target_cell
            )
        ):

            stats[
                "公式人工确认"
            ] += 1

            manual_result = (
                make_special_manual_result(
                    worksheet=worksheet,
                    row_index=row_index,
                    record_id=record_id,
                    korean_text=cell_text(
                        source_cell
                    ),
                    existing_chinese=cell_text(
                        target_cell
                    ),
                    reason=(
                        "当前行韩文或中文单元格包含公式，"
                        "程序不对公式结果执行自动严格审校。"
                    ),
                )
            )

            details.append(
                manual_result
            )

            stats[
                "人工确认"
            ] += 1

            continue


        # =================================================
        # 韩文
        # =================================================

        korean_text = cell_text(
            source_cell
        )


        if not korean_text.strip():

            stats[
                "空原文跳过"
            ] += 1

            continue


        if not contains_korean(
            korean_text
        ):

            stats[
                "非韩文跳过"
            ] += 1

            continue


        stats[
            "韩文有效行"
        ] += 1


        # =================================================
        # 现有中文
        # =================================================

        existing_chinese = cell_text(
            target_cell
        )


        if not existing_chinese.strip():

            stats[
                "现有中文为空"
            ] += 1


        # =================================================
        # Excel上下文
        # =================================================

        extra_context = (
            build_excel_context(
                worksheet=worksheet,
                row_index=row_index,
                source_column=source_column,
                header_map=header_map,
            )
        )


        # =================================================
        # 单条严格审校
        # =================================================

        result = review_text(
            korean_text=korean_text,
            existing_chinese=existing_chinese,
            knowledge_base=knowledge_base,
            extra_context=extra_context,
            record_id=record_id,
        )


        # =================================================
        # 致命错误立即停止整批
        # =================================================

        raise_if_fatal_review_error(
            result=result,
            location=(
                f"Excel工作表“{worksheet.title}”"
                f"第{row_index}行"
            ),
        )


        # =================================================
        # 审校明细
        # =================================================

        detail = {

            "工作表":
                worksheet.title,

            "原表行号":
                row_index,

            **result,
        }

        details.append(
            detail
        )


        update_review_stats(
            stats=stats,
            result=result,
        )


    # =====================================================
    # 全部处理完成后才生成报告
    # =====================================================

    add_review_reports(
        workbook=workbook,
        stats=stats,
        details=details,
        config=config,
    )


    # =====================================================
    # 最后才保存
    #
    # 中途如果Model API致命失败，
    # 不会运行到这里。
    # =====================================================

    output = BytesIO()

    workbook.save(
        output
    )

    output.seek(
        0
    )


    return (
        output.getvalue(),
        stats,
        details,
        config,
    )


# =========================================================
# 独立测试
# =========================================================

if __name__ == "__main__":

    print(
        "XLSX批量严格审校模块加载成功。"
    )