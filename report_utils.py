from io import BytesIO

from openpyxl import load_workbook


def analyze_confidence_levels(
    excel_bytes,
    sheet_name
):
    """
    从批量分析完成后的Excel中，
    统计 A / B / C / D / E 各可信等级数量。
    """

    workbook = load_workbook(
        BytesIO(excel_bytes),
        read_only=True,
        data_only=True
    )

    sheet = workbook[sheet_name]

    # =====================================================
    # 查找“检索_可信等级”所在列
    # =====================================================

    confidence_column = None

    for column in range(
        1,
        sheet.max_column + 1
    ):

        header = sheet.cell(
            row=1,
            column=column
        ).value

        if header == "检索_可信等级":

            confidence_column = column
            break


    if confidence_column is None:

        workbook.close()

        return {
            "A级": 0,
            "B级": 0,
            "C级": 0,
            "D级": 0,
            "E级": 0,
            "角色名专项": 0,
            "其他": 0,
            "总计": 0
        }


    stats = {
        "A级": 0,
        "B级": 0,
        "C级": 0,
        "D级": 0,
        "E级": 0,

        # 例如：
        # A级角色名／无整句译文
        "角色名专项": 0,

        "其他": 0,
        "总计": 0
    }


    # =====================================================
    # 逐行统计
    # =====================================================

    for row in range(
        2,
        sheet.max_row + 1
    ):

        value = sheet.cell(
            row=row,
            column=confidence_column
        ).value

        if value is None:
            continue

        value = str(value).strip()

        if not value:
            continue


        stats["总计"] += 1


        if value == "A级":

            stats["A级"] += 1


        elif value == "B级":

            stats["B级"] += 1


        elif value == "C级":

            stats["C级"] += 1


        elif value == "D级":

            stats["D级"] += 1


        elif value == "E级":

            stats["E级"] += 1


        elif value.startswith(
            "A级角色名"
        ):

            stats[
                "角色名专项"
            ] += 1


        else:

            stats["其他"] += 1


    workbook.close()

    return stats


def calculate_rates(
    confidence_stats
):
    """
    根据可信等级统计计算比例。
    """

    total = confidence_stats.get(
        "总计",
        0
    )

    if total <= 0:

        return {
            "A级率": 0.0,
            "历史资产覆盖率": 0.0,
            "需人工判断率": 0.0,
            "完全未命中率": 0.0
        }


    a = confidence_stats.get(
        "A级",
        0
    )

    b = confidence_stats.get(
        "B级",
        0
    )

    c = confidence_stats.get(
        "C级",
        0
    )

    d = confidence_stats.get(
        "D级",
        0
    )

    e = confidence_stats.get(
        "E级",
        0
    )

    role_only = confidence_stats.get(
        "角色名专项",
        0
    )


    return {

        # 可以直接复用整句历史译文
        "A级率":
            a / total * 100,

        # 公司历史翻译资产至少能够提供某种帮助
        "历史资产覆盖率":
            (
                a
                + b
                + c
                + d
                + role_only
            )
            /
            total
            *
            100,

        # 需要翻译人员判断
        "需人工判断率":
            (
                b
                + c
                + d
                + role_only
            )
            /
            total
            *
            100,

        # 三个正式库都没有足够参考
        "完全未命中率":
            e / total * 100
    }