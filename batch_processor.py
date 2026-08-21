from io import BytesIO
from copy import copy
import re

from openpyxl import load_workbook

from review_export import add_review_reports


# =========================================================
# 基础设置
# =========================================================

HANGUL_PATTERN = re.compile(r"[가-힣ㄱ-ㅎㅏ-ㅣ]")

CELL_TEXT_LIMIT = 30000

MAX_CONTEXT_RECORDS = 5


# =========================================================
# 韩文判断
# =========================================================

def contains_korean(value):
    """
    判断单元格是否包含韩文。
    """

    if value is None:
        return False

    return bool(
        HANGUL_PATTERN.search(
            str(value)
        )
    )


# =========================================================
# Excel单元格安全文本
# =========================================================

def safe_cell_text(value):
    """
    防止写入Excel的文本超过单元格限制。
    """

    if value is None:
        return ""

    text = str(value)

    if len(text) <= CELL_TEXT_LIMIT:
        return text

    return (
        text[:CELL_TEXT_LIMIT - 20]
        +
        "\n……【内容已截断】"
    )


# =========================================================
# Excel结构检查
# =========================================================

def inspect_excel(file_bytes):
    """
    检查上传Excel：

    - 工作表
    - 行数
    - 列数
    - 列名
    - 自动推荐韩文列
    """

    workbook = load_workbook(
        BytesIO(file_bytes),
        read_only=True,
        data_only=False
    )

    result = {}

    for sheet in workbook.worksheets:

        headers = []

        for column in range(
            1,
            sheet.max_column + 1
        ):

            value = sheet.cell(
                row=1,
                column=column
            ).value

            if value is None:

                headers.append(
                    f"未命名列_{column}"
                )

            else:

                headers.append(
                    str(value)
                )


        candidate_scores = []

        for column in range(
            1,
            sheet.max_column + 1
        ):

            header = headers[
                column - 1
            ]

            score = 0

            header_lower = (
                header
                .lower()
                .strip()
            )


            # =================================================
            # 常见韩文列名称
            # =================================================

            if header_lower == "msgid":
                score += 100

            if "韩文" in header:
                score += 80

            if "韩国语" in header:
                score += 80

            if "原文" in header:
                score += 30

            if "korean" in header_lower:
                score += 60

            if (
                header_lower
                ==
                "msgid_plural"
            ):
                score -= 100


            # =================================================
            # 抽样检测韩文比例
            # =================================================

            sample_count = 0
            korean_count = 0

            max_sample_row = min(
                sheet.max_row,
                101
            )

            for row in range(
                2,
                max_sample_row + 1
            ):

                value = sheet.cell(
                    row=row,
                    column=column
                ).value

                if value is None:
                    continue

                sample_count += 1

                if contains_korean(
                    value
                ):
                    korean_count += 1


            if sample_count > 0:

                korean_ratio = (
                    korean_count
                    /
                    sample_count
                )

                score += int(
                    korean_ratio * 50
                )


            candidate_scores.append(
                {
                    "列号": column,
                    "列名": header,
                    "得分": score
                }
            )


        candidate_scores.sort(
            key=lambda item:
                item["得分"],
            reverse=True
        )


        result[
            sheet.title
        ] = {
            "行数":
                sheet.max_row,

            "列数":
                sheet.max_column,

            "列名":
                headers,

            "韩文列候选":
                candidate_scores
        }


    workbook.close()

    return result


# =========================================================
# 提取不重复中文译文
# =========================================================

def unique_translations(records):
    """
    从历史记录中提取不重复中文译文。
    """

    translations = []

    for record in records:

        chinese = str(
            record.get(
                "msgstr[0]",
                ""
            )
        ).strip()

        if (
            chinese
            and
            chinese not in translations
        ):

            translations.append(
                chinese
            )

    return translations


# =========================================================
# 格式化长术语结果
# =========================================================

def format_long_terms(terms):
    """
    把长术语结果整理成适合Excel显示的文本。
    """

    if not terms:
        return ""

    lines = []

    for term in terms:

        korean_term = str(
            term.get(
                "韩文术语",
                ""
            )
        ).strip()

        records = term.get(
            "历史记录",
            []
        )

        translations = (
            unique_translations(
                records
            )
        )

        if translations:

            chinese_text = (
                " / ".join(
                    translations
                )
            )

            line = (
                f"{korean_term}"
                f" → "
                f"{chinese_text}"
            )

        else:

            line = korean_term


        if (
            line
            and
            line not in lines
        ):

            lines.append(
                line
            )


    return safe_cell_text(
        "\n".join(
            lines
        )
    )


# =========================================================
# 格式化历史上下文
# =========================================================

def format_context_matches(
    context_result
):
    """
    历史上下文最多写入前5条。
    """

    if not context_result:
        return ""

    results = context_result.get(
        "结果",
        []
    )

    if not results:
        return ""

    lines = []


    for item in results[
        :MAX_CONTEXT_RECORDS
    ]:

        historical_korean = str(
            item.get(
                "历史韩文",
                ""
            )
        ).strip()

        records = item.get(
            "历史记录",
            []
        )

        translations = (
            unique_translations(
                records
            )
        )


        if translations:

            chinese_text = (
                " / ".join(
                    translations
                )
            )

            block = (
                f"{historical_korean}"
                f"\n→ "
                f"{chinese_text}"
            )

        else:

            block = (
                historical_korean
            )


        if block:

            lines.append(
                block
            )


    total = context_result.get(
        "总数",
        len(results)
    )


    if (
        total
        >
        MAX_CONTEXT_RECORDS
    ):

        lines.append(
            f"……共找到 "
            f"{total} 条，"
            f"这里只显示前 "
            f"{MAX_CONTEXT_RECORDS} 条"
        )


    return safe_cell_text(
        "\n\n".join(
            lines
        )
    )


# =========================================================
# Excel批量处理
# =========================================================

def process_excel(
    file_bytes,
    sheet_name,
    korean_column,
    knowledge_base
):
    """
    批量检索：

    第一层：
    - 正式角色名
    - UWO完整句
    - Quest完整句

    第二层：
    只有完整句没有命中时执行：
    - UWO长术语
    - Quest长术语
    - UWO历史上下文
    - Quest历史上下文

    最后自动生成：

    - 待人工处理
    - D级_冲突记录
    - E级_完全未命中
    - 分析摘要
    """

    workbook = load_workbook(
        BytesIO(file_bytes),
        data_only=False
    )

    sheet = workbook[
        sheet_name
    ]


    # =====================================================
    # 新增分析列
    # =====================================================

    result_headers = [
        "检索_正式角色名",
        "检索_UWO精确译文",
        "检索_Quest精确译文",

        "检索_UWO长术语",
        "检索_Quest长术语",

        "检索_UWO历史上下文",
        "检索_Quest历史上下文",

        "检索_推荐中文",
        "检索_参考知识库",

        "检索_可信等级",
        "检索_匹配状态",
        "检索_深度匹配状态",

        "检索_需人工确认",
        "检索_说明"
    ]


    original_max_column = (
        sheet.max_column
    )

    start_column = (
        original_max_column + 1
    )


    # =====================================================
    # 继承标题格式
    # =====================================================

    source_header_cell = (
        sheet.cell(
            row=1,
            column=max(
                1,
                original_max_column
            )
        )
    )


    for offset, header in enumerate(
        result_headers
    ):

        cell = sheet.cell(
            row=1,
            column=
                start_column + offset
        )

        cell.value = header


        if source_header_cell.has_style:

            cell._style = copy(
                source_header_cell._style
            )


        cell.font = copy(
            source_header_cell.font
        )

        cell.fill = copy(
            source_header_cell.fill
        )

        cell.alignment = copy(
            source_header_cell.alignment
        )

        cell.border = copy(
            source_header_cell.border
        )

        cell.protection = copy(
            source_header_cell.protection
        )


    # =====================================================
    # 统计
    # =====================================================

    stats = {
        "总数据行":
            max(
                0,
                sheet.max_row - 1
            ),

        "实际处理": 0,
        "空韩文": 0,

        "角色名命中": 0,

        "UWO精确命中": 0,
        "Quest精确命中": 0,

        "高可信推荐": 0,

        "多译法或冲突": 0,

        "进入深度分析": 0,

        "UWO长术语命中": 0,
        "Quest长术语命中": 0,

        "UWO历史上下文命中": 0,
        "Quest历史上下文命中": 0,

        "深度参考命中": 0,

        "未命中": 0,
    }


    # =====================================================
    # 深度搜索缓存
    # =====================================================

    deep_cache = {}


    # =====================================================
    # 逐行处理
    # =====================================================

    for row_number in range(
        2,
        sheet.max_row + 1
    ):

        original_value = sheet.cell(
            row=row_number,
            column=korean_column
        ).value


        if original_value is None:

            stats[
                "空韩文"
            ] += 1

            continue


        korean_text = str(
            original_value
        ).strip()


        if not korean_text:

            stats[
                "空韩文"
            ] += 1

            continue


        stats[
            "实际处理"
        ] += 1


        # =================================================
        # 正式角色名
        # =================================================

        roles = (
            knowledge_base
            .find_roles_in_text(
                korean_text
            )
        )


        role_text = ""


        if roles:

            stats[
                "角色名命中"
            ] += 1


            role_text = "；".join(
                (
                    f"{item['韩文角色名']}"
                    f"→"
                    f"{item['正式中文名']}"
                )

                for item in roles
            )


        # =================================================
        # UWO完整句
        # =================================================

        uwo_records = (
            knowledge_base
            .search_uwo_exact(
                korean_text
            )
        )


        uwo_translations = (
            unique_translations(
                uwo_records
            )
        )


        if uwo_records:

            stats[
                "UWO精确命中"
            ] += 1


        # =================================================
        # Quest完整句
        # =================================================

        quest_records = (
            knowledge_base
            .search_quest_exact(
                korean_text
            )
        )


        quest_translations = (
            unique_translations(
                quest_records
            )
        )


        if quest_records:

            stats[
                "Quest精确命中"
            ] += 1


        # =================================================
        # 深度结果初始化
        # =================================================

        uwo_terms = []
        quest_terms = []


        uwo_context = {
            "总数": 0,
            "结果": [],
            "是否截断": False
        }


        quest_context = {
            "总数": 0,
            "结果": [],
            "是否截断": False
        }


        # =================================================
        # 没有完整句才进入深度分析
        # =================================================

        if (
            not uwo_translations
            and
            not quest_translations
        ):

            stats[
                "进入深度分析"
            ] += 1


            korean_character_count = len(
                HANGUL_PATTERN.findall(
                    korean_text
                )
            )


            if korean_character_count >= 2:

                if (
                    korean_text
                    in deep_cache
                ):

                    deep_result = (
                        deep_cache[
                            korean_text
                        ]
                    )

                else:

                    deep_result = (
                        knowledge_base
                        .search(
                            korean_text
                        )
                    )

                    deep_cache[
                        korean_text
                    ] = deep_result


                uwo_terms = (
                    deep_result.get(
                        "UWO长术语",
                        []
                    )
                )


                quest_terms = (
                    deep_result.get(
                        "Quest长术语",
                        []
                    )
                )


                uwo_context = (
                    deep_result.get(
                        "UWO包含匹配",
                        {
                            "总数": 0,
                            "结果": [],
                            "是否截断": False
                        }
                    )
                )


                quest_context = (
                    deep_result.get(
                        "Quest包含匹配",
                        {
                            "总数": 0,
                            "结果": [],
                            "是否截断": False
                        }
                    )
                )


        # =================================================
        # 深度结果格式化
        # =================================================

        uwo_terms_text = (
            format_long_terms(
                uwo_terms
            )
        )


        quest_terms_text = (
            format_long_terms(
                quest_terms
            )
        )


        uwo_context_text = (
            format_context_matches(
                uwo_context
            )
        )


        quest_context_text = (
            format_context_matches(
                quest_context
            )
        )


        # =================================================
        # 深度命中统计
        # =================================================

        if uwo_terms:

            stats[
                "UWO长术语命中"
            ] += 1


        if quest_terms:

            stats[
                "Quest长术语命中"
            ] += 1


        if (
            uwo_context.get(
                "总数",
                0
            )
            >
            0
        ):

            stats[
                "UWO历史上下文命中"
            ] += 1


        if (
            quest_context.get(
                "总数",
                0
            )
            >
            0
        ):

            stats[
                "Quest历史上下文命中"
            ] += 1


        # =================================================
        # 最终判断初始化
        # =================================================

        recommended = ""
        reference = ""

        confidence = ""
        match_status = ""
        deep_status = ""

        need_confirm = ""
        explanation = ""


        uwo_set = set(
            uwo_translations
        )


        quest_set = set(
            quest_translations
        )


        # =================================================
        # UWO完整句
        # =================================================

        if uwo_translations:

            reference = "UWO"


            if len(
                uwo_translations
            ) > 1:

                confidence = "D级"

                match_status = (
                    "UWO精确匹配-多译法"
                )

                deep_status = (
                    "未执行深度检索"
                )

                need_confirm = "是"

                explanation = (
                    "同一句韩文在UWO正式主库中"
                    "存在多个历史中文译法。"
                    "不能自动选择，"
                    "需结合Content、references"
                    "及实际场景判断。"
                )

                stats[
                    "多译法或冲突"
                ] += 1


            else:

                uwo_translation = (
                    uwo_translations[0]
                )


                if (
                    quest_translations
                    and
                    quest_set
                    !=
                    uwo_set
                ):

                    confidence = "D级"

                    match_status = (
                        "UWO/Quest精确匹配冲突"
                    )

                    deep_status = (
                        "未执行深度检索"
                    )

                    need_confirm = "是"

                    reference = (
                        "UWO；Quest"
                    )

                    explanation = (
                        "UWO和Quest均存在完整句"
                        "精确记录，但中文译法不一致。"
                        "需结合Content、script、"
                        "references和实际场景判断。"
                    )

                    stats[
                        "多译法或冲突"
                    ] += 1


                else:

                    confidence = "A级"

                    recommended = (
                        uwo_translation
                    )

                    match_status = (
                        "完整句精确匹配"
                    )

                    deep_status = (
                        "无需深度检索"
                    )

                    need_confirm = "否"

                    explanation = (
                        "UWO正式主库存在唯一"
                        "完整句精确译文，"
                        "可作为高可信历史译文复用。"
                    )

                    stats[
                        "高可信推荐"
                    ] += 1


        # =================================================
        # Quest完整句
        # =================================================

        elif quest_translations:

            reference = "Quest"


            if len(
                quest_translations
            ) > 1:

                confidence = "D级"

                match_status = (
                    "Quest精确匹配-多译法"
                )

                deep_status = (
                    "未执行深度检索"
                )

                need_confirm = "是"

                explanation = (
                    "同一句韩文在Quest正式主库中"
                    "存在多个历史中文译法。"
                    "需结合Content、script、"
                    "references和上下文判断。"
                )

                stats[
                    "多译法或冲突"
                ] += 1


            else:

                confidence = "A级"

                recommended = (
                    quest_translations[0]
                )

                match_status = (
                    "完整句精确匹配"
                )

                deep_status = (
                    "无需深度检索"
                )

                need_confirm = "否"

                explanation = (
                    "Quest正式主库存在唯一"
                    "完整句精确译文，"
                    "可作为高可信历史译文复用。"
                )

                stats[
                    "高可信推荐"
                ] += 1


        # =================================================
        # B级：长术语
        # =================================================

        elif (
            uwo_terms
            or
            quest_terms
        ):

            confidence = "B级"

            match_status = (
                "长术语历史译法匹配"
            )

            deep_status = (
                "深度检索已命中"
            )

            need_confirm = "是"

            references = []


            if uwo_terms:

                references.append(
                    "UWO长术语"
                )


            if quest_terms:

                references.append(
                    "Quest长术语"
                )


            if roles:

                references.insert(
                    0,
                    "角色名库"
                )


            reference = "；".join(
                references
            )


            explanation = (
                "没有完整句精确译文，"
                "但在正式主库历史文本中"
                "发现可复用的长术语。"
                "这些术语可以作为人工翻译的重要依据，"
                "但不能直接视为整句正式译文。"
            )


            stats[
                "深度参考命中"
            ] += 1


        # =================================================
        # C级：历史上下文
        # =================================================

        elif (
            uwo_context.get(
                "总数",
                0
            ) > 0
            or
            quest_context.get(
                "总数",
                0
            ) > 0
        ):

            confidence = "C级"

            match_status = (
                "历史上下文包含匹配"
            )

            deep_status = (
                "深度检索已命中"
            )

            need_confirm = "是"

            references = []


            if (
                uwo_context.get(
                    "总数",
                    0
                )
                >
                0
            ):

                references.append(
                    "UWO历史上下文"
                )


            if (
                quest_context.get(
                    "总数",
                    0
                )
                >
                0
            ):

                references.append(
                    "Quest历史上下文"
                )


            if roles:

                references.insert(
                    0,
                    "角色名库"
                )


            reference = "；".join(
                references
            )


            explanation = (
                "没有完整句或长术语精确记录，"
                "但发现历史长文本包含当前韩文。"
                "该结果仅用于人工判断上下文和历史译法，"
                "不能作为正式术语精确命中。"
            )


            stats[
                "深度参考命中"
            ] += 1


        # =================================================
        # 只有角色名
        # =================================================

        elif roles:

            confidence = (
                "A级角色名／无整句译文"
            )

            reference = (
                "角色名库"
            )

            match_status = (
                "仅正式角色名命中"
            )

            deep_status = (
                "深度检索未发现整句参考"
            )

            need_confirm = "是"

            explanation = (
                "正式角色名已确认，"
                "但UWO和Quest没有发现"
                "完整句、长术语或有效历史上下文。"
                "角色名必须使用正式中文名，"
                "整句仍需人工翻译。"
            )


        # =================================================
        # E级
        # =================================================

        else:

            confidence = "E级"

            reference = ""

            match_status = (
                "完全未命中"
            )

            deep_status = (
                "深度检索未命中"
            )

            need_confirm = "是"

            explanation = (
                "角色名库、UWO和Quest中"
                "均未发现足够可靠的历史记录。"
                "建议人工翻译或后续建立新术语候选。"
            )

            stats[
                "未命中"
            ] += 1


        # =================================================
        # 写入原工作表右侧
        # =================================================

        values = [

            safe_cell_text(
                role_text
            ),

            safe_cell_text(
                " || ".join(
                    uwo_translations
                )
            ),

            safe_cell_text(
                " || ".join(
                    quest_translations
                )
            ),

            uwo_terms_text,

            quest_terms_text,

            uwo_context_text,

            quest_context_text,

            safe_cell_text(
                recommended
            ),

            safe_cell_text(
                reference
            ),

            confidence,

            match_status,

            deep_status,

            need_confirm,

            safe_cell_text(
                explanation
            )
        ]


        for offset, value in enumerate(
            values
        ):

            sheet.cell(
                row=row_number,
                column=
                    start_column + offset
            ).value = value


    # =====================================================
    # 新增：生成待人工处理报告
    # =====================================================

    review_summary = (
        add_review_reports(
            workbook=
                workbook,

            source_sheet_name=
                sheet_name,

            korean_column=
                korean_column,

            batch_stats=
                stats
        )
    )


    # 把人工处理统计也放入stats，
    # 以后网页可以直接使用
    stats[
        "待人工处理"
    ] = review_summary.get(
        "待人工处理",
        0
    )

    stats[
        "D级冲突"
    ] = review_summary.get(
        "D级冲突",
        0
    )

    stats[
        "E级未命中"
    ] = review_summary.get(
        "E级未命中",
        0
    )


    # =====================================================
    # 保存新Excel
    # =====================================================

    output = BytesIO()

    workbook.save(
        output
    )

    workbook.close()

    output.seek(0)

    return (
        output.getvalue(),
        stats
    )