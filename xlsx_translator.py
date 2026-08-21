import re
from copy import copy
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from document_processor import (
    contains_korean,
    translate_text_unit,
    raise_if_fatal_ai_error,
    choose_exact_translation,
)


# =========================================================
# 韩中游戏本地化 Agent
# XLSX智能翻译
#
# 核心原则：
#
# 1. 不覆盖原文件
# 2. 不修改韩文原文
# 3. 不删除/合并/拆分/重排行
# 4. 已有非空中文默认保留
# 5. 没有中文列时新建 AI中文译文
# 6. 所有正式翻译统一走 translate_text_unit()
# 7. 最终QA结果直接使用 document_processor 的结果
# 8. 致命Model API错误立即停止整批
# 9. 不生成部分完成的输出文件
#
#
# 翻译链路：
#
# XLSX
# ↓
# translate_text_unit
# ↓
# 知识库
# ↓
# model_gateway
# ↓
# 换行保护
# ↓
# 标点自动规范化
# ↓
# 韩文残留/数字/标签/占位符等最终QA
# ↓
# XLSX写入
# ↓
# 审计报告
#
# =========================================================


# =========================================================
# 自动生成的报告工作表
# =========================================================

GENERATED_SHEETS = [
    "AI翻译处理摘要",
    "AI翻译明细",
    "需人工确认",
    "格式异常",
]


# =========================================================
# 字段关键词
# =========================================================

KOREAN_HEADER_EXACT = {
    "韩文",
    "韩文原文",
    "韩语",
    "韩语原文",
    "korean",
    "한국어",
    "원문",
    "msgid",
}


CHINESE_HEADER_EXACT = {
    "中文",
    "中文译文",
    "中文翻译",
    "简体中文",
    "译文",
    "msgstr",
    "msgstr[0]",
    "ai中文译文",
    "ai建议译文",
}


METADATA_ALIASES = {
    "ID": [
        "id",
        "stringid",
        "string_id",
        "key",
        "文本id",
        "字符串id",
    ],

    "msgctxt": [
        "msgctxt",
        "context",
        "上下文",
    ],

    "references": [
        "references",
        "reference",
        "refs",
        "引用",
    ],

    "Release": [
        "release",
        "版本",
    ],

    "Content": [
        "content",
        "内容",
    ],

    "script": [
        "script",
        "脚本",
    ],
}


CHINESE_PATTERN = re.compile(
    r"[\u3400-\u4DBF\u4E00-\u9FFF]"
)


MANUAL_MARKER_PATTERN = re.compile(
    r"【需人工确认\s*[:：]\s*(.*?)】",
    flags=re.DOTALL,
)


# =========================================================
# 自定义异常
# =========================================================

class SheetSelectionRequired(
    Exception
):
    pass


class ColumnSelectionRequired(
    Exception
):
    pass


# =========================================================
# 基础文本
# =========================================================

def normalize_header(
    value
):

    if value is None:
        return ""


    text = str(
        value
    ).strip().lower()


    text = re.sub(
        r"\s+",
        "",
        text,
    )


    return text


def cell_text(
    cell
):

    if cell is None:
        return ""


    value = cell.value


    if value is None:
        return ""


    return str(
        value
    ).strip()


def is_formula_cell(
    cell
):

    if cell is None:
        return False


    value = cell.value


    return (
        isinstance(
            value,
            str,
        )
        and
        value.startswith(
            "="
        )
    )


# =========================================================
# 表头检测
# =========================================================

def detect_header_row(
    worksheet,
    max_scan_rows=30,
):

    scan_end = min(
        worksheet.max_row,
        max_scan_rows,
    )


    best_row = 1
    best_score = -1


    known_headers = {
        normalize_header(
            item
        )
        for item
        in (
            KOREAN_HEADER_EXACT
            |
            CHINESE_HEADER_EXACT
        )
    }


    for aliases in (
        METADATA_ALIASES.values()
    ):

        for alias in aliases:

            known_headers.add(
                normalize_header(
                    alias
                )
            )


    for row_index in range(
        1,
        scan_end + 1,
    ):

        score = 0
        non_empty = 0


        for column_index in range(
            1,
            worksheet.max_column + 1,
        ):

            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value


            if value is None:
                continue


            text = str(
                value
            ).strip()


            if not text:
                continue


            non_empty += 1


            normalized = normalize_header(
                text
            )


            if normalized in known_headers:

                score += 40


            if normalized in {
                "msgid",
                "韩文",
                "韩文原文",
                "한국어",
                "원문",
            }:

                score += 80


            if normalized in {
                "msgstr[0]",
                "中文",
                "中文译文",
                "简体中文",
            }:

                score += 60


        score += min(
            non_empty,
            15,
        )


        if score > best_score:

            best_score = score
            best_row = row_index


    return best_row


# =========================================================
# 表头列表
# =========================================================

def get_headers(
    worksheet,
    header_row,
):

    result = []


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        raw_value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        header_name = str(
            raw_value
            or
            ""
        ).strip()


        if not header_name:

            header_name = (
                f"未命名列{column_index}"
            )


        result.append(
            {
                "列号":
                    column_index,

                "列字母":
                    get_column_letter(
                        column_index
                    ),

                "列名":
                    header_name,
            }
        )


    return result


# =========================================================
# 列采样
# =========================================================

def sample_column_values(
    worksheet,
    column_index,
    header_row,
    sample_size=30,
):

    result = []


    row_end = min(
        worksheet.max_row,
        header_row
        +
        sample_size
        +
        10,
    )


    for row_index in range(
        header_row + 1,
        row_end + 1,
    ):

        value = worksheet.cell(
            row=row_index,
            column=column_index,
        ).value


        if value is None:

            continue


        if isinstance(
            value,
            str,
        ):

            text = value.strip()

        else:

            text = str(
                value
            ).strip()


        if not text:

            continue


        result.append(
            text
        )


        if len(
            result
        ) >= sample_size:

            break


    return result


# =========================================================
# 韩文列评分
#
# 常见情况：
#
# 韩文表头 + 韩文数据
# → 250分左右
# =========================================================

def score_korean_column(
    worksheet,
    column_index,
    header_row,
):

    header = normalize_header(
        worksheet.cell(
            row=header_row,
            column=column_index,
        ).value
    )


    score = 0


    if header in {
        normalize_header(
            item
        )
        for item
        in KOREAN_HEADER_EXACT
    }:

        score += 200


    elif (
        "韩文" in header
        or
        "韩语" in header
        or
        "korean" in header
        or
        "한국" in header
        or
        "msgid" in header
    ):

        score += 150


    samples = sample_column_values(
        worksheet=
            worksheet,

        column_index=
            column_index,

        header_row=
            header_row,
    )


    if samples:

        korean_count = sum(
            1
            for text in samples
            if contains_korean(
                text
            )
        )


        ratio = (
            korean_count
            /
            len(
                samples
            )
        )


        score += round(
            ratio
            *
            50
        )


    return score


# =========================================================
# 中文列评分
#
# 常见情况：
#
# 中文表头 + 中文数据
# → 150分左右
# =========================================================

def score_chinese_column(
    worksheet,
    column_index,
    header_row,
):

    header = normalize_header(
        worksheet.cell(
            row=header_row,
            column=column_index,
        ).value
    )


    score = 0


    if header in {
        normalize_header(
            item
        )
        for item
        in CHINESE_HEADER_EXACT
    }:

        score += 120


    elif (
        "中文" in header
        or
        "译文" in header
        or
        "chinese" in header
        or
        "msgstr" in header
    ):

        score += 90


    samples = sample_column_values(
        worksheet=
            worksheet,

        column_index=
            column_index,

        header_row=
            header_row,
    )


    if samples:

        chinese_count = sum(
            1
            for text in samples
            if CHINESE_PATTERN.search(
                text
            )
        )


        ratio = (
            chinese_count
            /
            len(
                samples
            )
        )


        score += round(
            ratio
            *
            30
        )


    return score


# =========================================================
# 列候选
# =========================================================

def find_column_candidates(
    worksheet,
    header_row,
    column_type,
):

    candidates = []


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        if column_type == "韩文":

            score = score_korean_column(
                worksheet=
                    worksheet,

                column_index=
                    column_index,

                header_row=
                    header_row,
            )


        elif column_type == "中文":

            score = score_chinese_column(
                worksheet=
                    worksheet,

                column_index=
                    column_index,

                header_row=
                    header_row,
            )


        else:

            raise ValueError(
                f"未知列类型：{column_type}"
            )


        if score <= 0:

            continue


        header_value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        candidates.append(
            {
                "列号":
                    column_index,

                "列字母":
                    get_column_letter(
                        column_index
                    ),

                "列名":
                    str(
                        header_value
                        or
                        f"未命名列{column_index}"
                    ).strip(),

                "得分":
                    score,
            }
        )


    candidates.sort(
        key=lambda item:
            item.get(
                "得分",
                0,
            ),
        reverse=True,
    )


    return candidates


# =========================================================
# 推荐候选
# =========================================================

def recommend_candidate(
    candidates
):

    candidates = (
        candidates
        or
        []
    )


    if not candidates:

        return None


    first = candidates[
        0
    ]


    if len(
        candidates
    ) == 1:

        return first


    second = candidates[
        1
    ]


    if (
        first.get(
            "得分",
            0,
        )
        -
        second.get(
            "得分",
            0,
        )
        >=
        20
    ):

        return first


    return None


# =========================================================
# 列解析
#
# 支持：
#
# 2
# "2"
# "B"
# "中文"
# =========================================================

def resolve_column(
    worksheet,
    header_row,
    column_value,
):

    if column_value is None:

        return None


    # =====================================================
    # int
    # =====================================================

    if isinstance(
        column_value,
        int,
    ):

        if (
            1
            <=
            column_value
            <=
            worksheet.max_column
        ):

            return column_value


        raise ColumnSelectionRequired(
            (
                f"列号超出范围："
                f"{column_value}"
            )
        )


    text = str(
        column_value
    ).strip()


    if not text:

        return None


    # =====================================================
    # 数字字符串
    # =====================================================

    if text.isdigit():

        number = int(
            text
        )


        if (
            1
            <=
            number
            <=
            worksheet.max_column
        ):

            return number


    # =====================================================
    # Excel列字母
    # =====================================================

    if re.fullmatch(
        r"[A-Za-z]+",
        text,
    ):

        from openpyxl.utils.cell import (
            column_index_from_string,
        )


        try:

            number = (
                column_index_from_string(
                    text
                )
            )


            if (
                1
                <=
                number
                <=
                worksheet.max_column
            ):

                return number


        except Exception:

            pass


    # =====================================================
    # 表头名
    # =====================================================

    target = normalize_header(
        text
    )


    matches = []


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        header_value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        if normalize_header(
            header_value
        ) == target:

            matches.append(
                column_index
            )


    if len(
        matches
    ) == 1:

        return matches[
            0
        ]


    if len(
        matches
    ) > 1:

        raise ColumnSelectionRequired(
            (
                f"表头“{text}”出现多次，"
                "请人工选择具体列。"
            )
        )


    raise ColumnSelectionRequired(
        f"无法找到列：{column_value}"
    )


# =========================================================
# 工作表选择
# =========================================================

def choose_sheet(
    workbook,
    sheet_name=None,
):

    if sheet_name:

        if sheet_name not in workbook.sheetnames:

            raise SheetSelectionRequired(
                (
                    f"工作表不存在："
                    f"{sheet_name}"
                )
            )


        return workbook[
            sheet_name
        ]


    business_sheets = [
        worksheet
        for worksheet
        in workbook.worksheets
        if worksheet.title
        not in GENERATED_SHEETS
    ]


    if len(
        business_sheets
    ) == 1:

        return business_sheets[
            0
        ]


    visible_sheets = [
        worksheet
        for worksheet
        in business_sheets
        if worksheet.sheet_state
        ==
        "visible"
    ]


    if len(
        visible_sheets
    ) == 1:

        return visible_sheets[
            0
        ]


    raise SheetSelectionRequired(
        (
            "Excel包含多个可处理工作表，"
            "请在界面中人工选择。"
        )
    )


# =========================================================
# 韩文列选择
# =========================================================

def choose_korean_column(
    worksheet,
    header_row,
    korean_column=None,
):

    if korean_column is not None:

        return resolve_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            column_value=
                korean_column,
        )


    candidates = (
        find_column_candidates(
            worksheet=
                worksheet,

            header_row=
                header_row,

            column_type=
                "韩文",
        )
    )


    recommended = (
        recommend_candidate(
            candidates
        )
    )


    if recommended:

        return recommended[
            "列号"
        ]


    raise ColumnSelectionRequired(
        (
            "无法唯一确定韩文原文列。"
            "请在界面中人工选择。"
        )
    )


# =========================================================
# 最后表头列
# =========================================================

def last_header_column(
    worksheet,
    header_row,
):

    last = 0


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        if (
            value is not None
            and
            str(
                value
            ).strip()
        ):

            last = column_index


    return (
        last
        if last > 0
        else
        worksheet.max_column
    )


# =========================================================
# 样式复制
# =========================================================

def copy_cell_style(
    source_cell,
    target_cell,
):

    if source_cell is None:
        return


    if source_cell.has_style:

        target_cell._style = copy(
            source_cell._style
        )


    if source_cell.number_format:

        target_cell.number_format = (
            source_cell.number_format
        )


    if source_cell.alignment:

        target_cell.alignment = copy(
            source_cell.alignment
        )


    if source_cell.protection:

        target_cell.protection = copy(
            source_cell.protection
        )


# =========================================================
# 新建AI中文列
# =========================================================

def create_ai_translation_column(
    worksheet,
    header_row,
    korean_column,
):

    new_column = (
        last_header_column(
            worksheet,
            header_row,
        )
        +
        1
    )


    header_cell = worksheet.cell(
        row=header_row,
        column=new_column,
    )


    header_cell.value = (
        "AI中文译文"
    )


    source_header = worksheet.cell(
        row=header_row,
        column=korean_column,
    )


    copy_cell_style(
        source_header,
        header_cell,
    )


    return new_column


# =========================================================
# 中文列选择
#
# 无候选：
# → 新建AI中文译文
#
# 多候选且不唯一：
# → 人工选择
# =========================================================

def choose_or_create_chinese_column(
    worksheet,
    header_row,
    korean_column,
    chinese_column=None,
):

    if chinese_column is not None:

        resolved = resolve_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            column_value=
                chinese_column,
        )


        if resolved == korean_column:

            raise ColumnSelectionRequired(
                "韩文列和中文列不能相同。"
            )


        return (
            resolved,
            False,
        )


    candidates = (
        find_column_candidates(
            worksheet=
                worksheet,

            header_row=
                header_row,

            column_type=
                "中文",
        )
    )


    candidates = [
        item
        for item in candidates
        if item.get(
            "列号"
        )
        !=
        korean_column
    ]


    if not candidates:

        created = (
            create_ai_translation_column(
                worksheet=
                    worksheet,

                header_row=
                    header_row,

                korean_column=
                    korean_column,
            )
        )


        return (
            created,
            True,
        )


    recommended = (
        recommend_candidate(
            candidates
        )
    )


    if recommended:

        return (
            recommended[
                "列号"
            ],
            False,
        )


    raise ColumnSelectionRequired(
        (
            "检测到多个可能的中文目标列，"
            "无法安全自动选择。"
            "请在界面中人工指定。"
        )
    )


# =========================================================
# 表头映射
# =========================================================

def build_header_map(
    worksheet,
    header_row,
):

    result = {}


    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        value = worksheet.cell(
            row=header_row,
            column=column_index,
        ).value


        normalized = normalize_header(
            value
        )


        if (
            normalized
            and
            normalized not in result
        ):

            result[
                normalized
            ] = column_index


    return result


# =========================================================
# 元数据值
# =========================================================

def find_metadata_value(
    worksheet,
    row_index,
    header_map,
    aliases,
):

    for alias in aliases:

        normalized = normalize_header(
            alias
        )


        column_index = (
            header_map.get(
                normalized
            )
        )


        if not column_index:

            continue


        value = worksheet.cell(
            row=row_index,
            column=column_index,
        ).value


        if value is None:
            continue


        text = str(
            value
        ).strip()


        if text:

            return text


    return ""


# =========================================================
# Excel上下文
#
# strict_batch_reviewer也会调用这个函数，
# 所以保持通用。
# =========================================================

def build_excel_context(
    worksheet,
    row_index,
    header_row,
    korean_column,
    header_map=None,
    chinese_column=None,
):

    if header_map is None:

        header_map = (
            build_header_map(
                worksheet,
                header_row,
            )
        )


    parts = []


    # =====================================================
    # 元数据
    # =====================================================

    for field_name in [
        "ID",
        "msgctxt",
        "references",
        "Release",
        "Content",
        "script",
    ]:

        value = (
            find_metadata_value(
                worksheet=
                    worksheet,

                row_index=
                    row_index,

                header_map=
                    header_map,

                aliases=
                    METADATA_ALIASES.get(
                        field_name,
                        [],
                    ),
            )
        )


        if value:

            parts.append(
                f"{field_name}：{value}"
            )


    # =====================================================
    # 上一条韩文
    # =====================================================

    if row_index > header_row + 1:

        previous_value = (
            worksheet.cell(
                row=row_index - 1,
                column=korean_column,
            ).value
        )


        if previous_value is not None:

            previous_text = str(
                previous_value
            ).strip()


            if previous_text:

                parts.append(
                    (
                        "上一条韩文："
                        +
                        previous_text
                    )
                )


    # =====================================================
    # 下一条韩文
    # =====================================================

    if row_index < worksheet.max_row:

        next_value = (
            worksheet.cell(
                row=row_index + 1,
                column=korean_column,
            ).value
        )


        if next_value is not None:

            next_text = str(
                next_value
            ).strip()


            if next_text:

                parts.append(
                    (
                        "下一条韩文："
                        +
                        next_text
                    )
                )


    return "\n".join(
        parts
    )


# =========================================================
# 人工确认标记清理
#
# 主要保留给旧代码和回归测试。
#
# translate_text_unit现在已经会提前清理。
# =========================================================

def strip_manual_marker(
    translation
):

    translation = str(
        translation
        or
        ""
    )


    matches = (
        MANUAL_MARKER_PATTERN.findall(
            translation
        )
    )


    reasons = []


    for match in matches:

        reason = str(
            match
            or
            ""
        ).strip()


        if (
            reason
            and
            reason not in reasons
        ):

            reasons.append(
                reason
            )


    cleaned = (
        MANUAL_MARKER_PATTERN.sub(
            "",
            translation,
        ).rstrip()
    )


    return (
        cleaned,
        "；".join(
            reasons
        ),
    )


# =========================================================
# XLSX统计
# =========================================================

def new_translation_stats():

    return {
        "总数据行":
            0,

        "韩文文本数":
            0,

        "已有中文保留":
            0,

        "新增译文":
            0,

        "正式完整句直接复用":
            0,

        "AI翻译成功":
            0,

        "AI翻译失败":
            0,

        "需人工确认":
            0,

        "格式异常":
            0,

        "标点自动规范化":
            0,

        "公式跳过":
            0,

        "合并单元格跳过":
            0,

        "非韩文跳过":
            0,

        "空文本跳过":
            0,
    }


# =========================================================
# 更新统计
# =========================================================

def update_translation_stats(
    stats,
    result,
):

    method = str(
        result.get(
            "处理方式",
            ""
        )
        or
        ""
    )


    if method in {
        "空文本跳过",
        "纯空白跳过",
    }:

        stats[
            "空文本跳过"
        ] += 1

        return


    if method == "非韩文跳过":

        stats[
            "非韩文跳过"
        ] += 1

        return


    stats[
        "韩文文本数"
    ] += 1


    if method == "正式完整句直接复用":

        stats[
            "正式完整句直接复用"
        ] += 1


    if result.get(
        "AI成功",
        False,
    ):

        stats[
            "AI翻译成功"
        ] += 1


    if (
        method
        ==
        "AI调用失败，保留韩文"
    ):

        stats[
            "AI翻译失败"
        ] += 1


    if result.get(
        "需人工确认",
        False,
    ):

        stats[
            "需人工确认"
        ] += 1


    if result.get(
        "最终QA通过"
    ) is False:

        stats[
            "格式异常"
        ] += 1


    if result.get(
        "标点自动规范化",
        False,
    ):

        stats[
            "标点自动规范化"
        ] += 1


# =========================================================
# 删除旧报告
# =========================================================

def remove_old_generated_sheets(
    workbook
):

    for sheet_name in GENERATED_SHEETS:

        if sheet_name in workbook.sheetnames:

            workbook.remove(
                workbook[
                    sheet_name
                ]
            )


# =========================================================
# 报告格式
# =========================================================

def format_report_sheet(
    worksheet
):

    if worksheet.max_row < 1:

        return


    # =====================================================
    # 表头
    # =====================================================

    for cell in worksheet[
        1
    ]:

        cell.font = Font(
            bold=True
        )


        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )


        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )


    worksheet.freeze_panes = (
        "A2"
    )


    # =====================================================
    # 正文换行
    # =====================================================

    for row in worksheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


    # =====================================================
    # 宽度
    # =====================================================

    for column_index in range(
        1,
        worksheet.max_column + 1,
    ):

        max_length = 0


        for row_index in range(
            1,
            min(
                worksheet.max_row,
                200,
            )
            + 1,
        ):

            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value


            if value is None:

                continue


            text = str(
                value
            )


            current_length = min(
                len(
                    text
                ),
                80,
            )


            max_length = max(
                max_length,
                current_length,
            )


        width = min(
            max(
                max_length
                +
                2,
                10,
            ),
            45,
        )


        worksheet.column_dimensions[
            get_column_letter(
                column_index
            )
        ].width = width


    if (
        worksheet.max_row >= 2
        and
        worksheet.max_column >= 1
    ):

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )


# =========================================================
# 字典列表写入
# =========================================================

def write_dict_rows(
    worksheet,
    rows
):

    rows = (
        rows
        or
        []
    )


    if not rows:

        worksheet.append(
            [
                "无记录"
            ]
        )

        return


    headers = []


    for row in rows:

        for key in row.keys():

            if key not in headers:

                headers.append(
                    key
                )


    worksheet.append(
        headers
    )


    for row in rows:

        worksheet.append(
            [
                row.get(
                    header,
                    "",
                )
                for header
                in headers
            ]
        )


# =========================================================
# 添加报告
# =========================================================

def add_translation_reports(
    workbook,
    stats,
    details,
    config,
):

    remove_old_generated_sheets(
        workbook
    )


    # =====================================================
    # 摘要
    # =====================================================

    summary = workbook.create_sheet(
        "AI翻译处理摘要"
    )


    summary.append(
        [
            "项目",
            "结果",
        ]
    )


    for key, value in config.items():

        summary.append(
            [
                key,
                value,
            ]
        )


    for key, value in stats.items():

        summary.append(
            [
                key,
                value,
            ]
        )


    format_report_sheet(
        summary
    )


    # =====================================================
    # 全部明细
    # =====================================================

    detail_sheet = workbook.create_sheet(
        "AI翻译明细"
    )


    write_dict_rows(
        detail_sheet,
        details,
    )


    format_report_sheet(
        detail_sheet
    )


    # =====================================================
    # 人工确认
    # =====================================================

    manual_rows = [
        item
        for item in details
        if item.get(
            "需人工确认",
            False,
        )
    ]


    manual_sheet = workbook.create_sheet(
        "需人工确认"
    )


    write_dict_rows(
        manual_sheet,
        manual_rows,
    )


    format_report_sheet(
        manual_sheet
    )


    # =====================================================
    # 最终QA异常
    #
    # 重要：
    #
    # 不再重新执行旧format_qa。
    #
    # 直接使用最终准备写入Excel的中文
    # 已经得到的“最终QA通过”字段。
    # =====================================================

    qa_rows = [
        item
        for item in details
        if item.get(
            "最终QA通过"
        )
        is False
    ]


    qa_sheet = workbook.create_sheet(
        "格式异常"
    )


    write_dict_rows(
        qa_sheet,
        qa_rows,
    )


    format_report_sheet(
        qa_sheet
    )


# =========================================================
# XLSX结构检查
# =========================================================

def inspect_xlsx_for_translation(
    file_bytes
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )


    result = {}


    for worksheet in workbook.worksheets:

        if worksheet.title in GENERATED_SHEETS:

            continue


        header_row = detect_header_row(
            worksheet
        )


        korean_candidates = (
            find_column_candidates(
                worksheet=
                    worksheet,

                header_row=
                    header_row,

                column_type=
                    "韩文",
            )
        )


        chinese_candidates = (
            find_column_candidates(
                worksheet=
                    worksheet,

                header_row=
                    header_row,

                column_type=
                    "中文",
            )
        )


        recommended_korean = (
            recommend_candidate(
                korean_candidates
            )
        )


        recommended_chinese = (
            recommend_candidate(
                chinese_candidates
            )
        )


        result[
            worksheet.title
        ] = {
            "工作表":
                worksheet.title,

            "隐藏状态":
                worksheet.sheet_state,

            "行数":
                worksheet.max_row,

            "列数":
                worksheet.max_column,

            "表头行":
                header_row,

            "列名":
                get_headers(
                    worksheet,
                    header_row,
                ),

            "韩文列候选":
                korean_candidates,

            "中文列候选":
                chinese_candidates,

            "推荐韩文列":
                (
                    recommended_korean[
                        "列号"
                    ]
                    if recommended_korean
                    else None
                ),

            "推荐中文列":
                (
                    recommended_chinese[
                        "列号"
                    ]
                    if recommended_chinese
                    else None
                ),
        }


    return result


# =========================================================
# 统一进度回调
#
# callback只接收一个dict，便于Streamlit或其他界面复用。
# 回调自身出错时不影响翻译核心流程。
# =========================================================

def emit_progress(
    progress_callback,
    current,
    total,
    stage,
    message="",
    row_index=None,
    stats=None,
):

    if not callable(
        progress_callback
    ):

        return


    total = max(
        int(
            total
            or
            0
        ),
        0,
    )


    current = max(
        int(
            current
            or
            0
        ),
        0,
    )


    if total > 0:

        percent = min(
            max(
                current
                /
                total,
                0.0,
            ),
            1.0,
        )

    else:

        percent = 1.0


    event = {
        "当前":
            current,

        "总数":
            total,

        "进度":
            percent,

        "阶段":
            str(
                stage
                or
                ""
            ),

        "消息":
            str(
                message
                or
                ""
            ),

        "行号":
            row_index,

        "统计":
            dict(
                stats
                or
                {}
            ),
    }


    try:

        progress_callback(
            event
        )

    except Exception:

        # UI进度显示失败绝不能影响核心翻译。
        pass


# =========================================================
# 预检阶段的中文目标列解析
#
# 与正式翻译不同：
# 预检绝不真正创建“AI中文译文”列。
# 没有中文列时返回None并标记后续将新建。
# =========================================================

def resolve_preflight_chinese_column(
    worksheet,
    header_row,
    korean_column,
    chinese_column=None,
):

    if chinese_column is not None:

        resolved = resolve_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            column_value=
                chinese_column,
        )


        if resolved == korean_column:

            raise ColumnSelectionRequired(
                "韩文列和中文列不能相同。"
            )


        return (
            resolved,
            False,
        )


    candidates = (
        find_column_candidates(
            worksheet=
                worksheet,

            header_row=
                header_row,

            column_type=
                "中文",
        )
    )


    candidates = [
        item
        for item in candidates
        if item.get(
            "列号"
        )
        !=
        korean_column
    ]


    if not candidates:

        return (
            None,
            True,
        )


    recommended = (
        recommend_candidate(
            candidates
        )
    )


    if recommended:

        return (
            recommended[
                "列号"
            ],
            False,
        )


    raise ColumnSelectionRequired(
        (
            "检测到多个可能的中文目标列，"
            "无法安全自动选择。"
            "请在界面中人工指定。"
        )
    )


# =========================================================
# Excel翻译前深度预检
#
# 特点：
#
# - 不调用Model API
# - 不消耗Token
# - 不修改Excel
# - 不创建新列
# - 使用本地知识库估算正式完整句可直接复用量
#
# 返回：
#
# preflight_stats, config
# =========================================================

def preflight_xlsx_translation(
    file_bytes,
    knowledge_base,
    sheet_name=None,
    korean_column=None,
    chinese_column=None,
    progress_callback=None,
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )


    worksheet = choose_sheet(
        workbook=
            workbook,

        sheet_name=
            sheet_name,
    )


    header_row = detect_header_row(
        worksheet
    )


    korean_column = (
        choose_korean_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            korean_column=
                korean_column,
        )
    )


    (
        resolved_chinese_column,
        will_create_chinese_column,
    ) = (
        resolve_preflight_chinese_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            korean_column=
                korean_column,

            chinese_column=
                chinese_column,
        )
    )


    header_map = build_header_map(
        worksheet=
            worksheet,

        header_row=
            header_row,
    )


    stats = {
        "总数据行":
            0,

        "韩文文本数":
            0,

        "已有中文保留":
            0,

        "待翻译":
            0,

        "正式完整句预计复用":
            0,

        "预计AI请求":
            0,

        "知识库风险":
            0,

        "公式跳过":
            0,

        "合并单元格跳过":
            0,

        "非韩文跳过":
            0,

        "空文本跳过":
            0,
    }


    total_rows = max(
        worksheet.max_row
        -
        header_row,
        0,
    )


    emit_progress(
        progress_callback=
            progress_callback,

        current=
            0,

        total=
            total_rows,

        stage=
            "预检",

        message=
            "正在扫描Excel并估算知识库复用率...",

        stats=
            stats,
    )


    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):

        current = (
            row_index
            -
            header_row
            -
            1
        )


        emit_progress(
            progress_callback=
                progress_callback,

            current=
                current,

            total=
                total_rows,

            stage=
                "预检",

            message=(
                f"正在预检第"
                f"{current + 1:,}"
                f"/"
                f"{total_rows:,}"
                f"行"
            ),

            row_index=
                row_index,

            stats=
                stats,
        )


        stats[
            "总数据行"
        ] += 1


        source_cell = worksheet.cell(
            row=
                row_index,

            column=
                korean_column,
        )


        target_cell = (
            worksheet.cell(
                row=
                    row_index,

                column=
                    resolved_chinese_column,
            )
            if resolved_chinese_column
            is not None
            else
            None
        )


        if isinstance(
            source_cell,
            MergedCell,
        ):

            stats[
                "合并单元格跳过"
            ] += 1

            continue


        if (
            target_cell is not None
            and
            isinstance(
                target_cell,
                MergedCell,
            )
        ):

            stats[
                "合并单元格跳过"
            ] += 1

            continue


        if (
            is_formula_cell(
                source_cell
            )
            or
            (
                target_cell
                is not None
                and
                is_formula_cell(
                    target_cell
                )
            )
        ):

            stats[
                "公式跳过"
            ] += 1

            continue


        korean_text = str(
            source_cell.value
            or
            ""
        )


        if not korean_text:

            stats[
                "空文本跳过"
            ] += 1

            continue


        if not contains_korean(
            korean_text
        ):

            stats[
                "非韩文跳过"
            ] += 1

            continue


        stats[
            "韩文文本数"
        ] += 1


        existing_chinese = (
            str(
                target_cell.value
                or
                ""
            ).strip()
            if target_cell
            is not None
            else
            ""
        )


        if existing_chinese:

            stats[
                "已有中文保留"
            ] += 1

            continue


        stats[
            "待翻译"
        ] += 1


        core_text = (
            korean_text.strip()
        )


        search_result = (
            knowledge_base.search(
                core_text
            )
        )


        exact_decision = (
            choose_exact_translation(
                search_result
            )
        )


        if exact_decision.get(
            "可直接复用",
            False,
        ):

            stats[
                "正式完整句预计复用"
            ] += 1

        else:

            stats[
                "预计AI请求"
            ] += 1


            if exact_decision.get(
                "存在风险",
                False,
            ):

                stats[
                    "知识库风险"
                ] += 1


    config = {
        "处理工作表":
            worksheet.title,

        "表头行":
            header_row,

        "韩文列":
            (
                f"{get_column_letter(korean_column)}"
                f" / "
                f"{worksheet.cell(header_row, korean_column).value}"
            ),

        "中文目标策略":
            (
                "新建 AI中文译文"
                if will_create_chinese_column
                else
                (
                    f"{get_column_letter(resolved_chinese_column)}"
                    f" / "
                    f"{worksheet.cell(header_row, resolved_chinese_column).value}"
                )
            ),

        "已有中文策略":
            "默认保留，不覆盖",

        "是否调用Model API":
            "否",

        "是否消耗Token":
            "否",

        "正式库复用估算":
            "基于当前本地角色名/UWO/Quest知识库",
    }


    emit_progress(
        progress_callback=
            progress_callback,

        current=
            total_rows,

        total=
            total_rows,

        stage=
            "预检完成",

        message=
            "Excel翻译前预检完成。",

        stats=
            stats,
    )


    return (
        stats,
        config,
    )


# =========================================================
# XLSX主翻译入口
# =========================================================

def process_xlsx_translation(
    file_bytes,
    knowledge_base,
    sheet_name=None,
    korean_column=None,
    chinese_column=None,
    progress_callback=None,
):

    workbook = load_workbook(
        BytesIO(
            file_bytes
        ),
        data_only=False,
    )


    # =====================================================
    # 选择工作表
    # =====================================================

    worksheet = choose_sheet(
        workbook=
            workbook,

        sheet_name=
            sheet_name,
    )


    header_row = detect_header_row(
        worksheet
    )


    # =====================================================
    # 韩文列
    # =====================================================

    korean_column = (
        choose_korean_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            korean_column=
                korean_column,
        )
    )


    # =====================================================
    # 中文列
    # =====================================================

    (
        chinese_column,
        created_chinese_column,
    ) = (
        choose_or_create_chinese_column(
            worksheet=
                worksheet,

            header_row=
                header_row,

            korean_column=
                korean_column,

            chinese_column=
                chinese_column,
        )
    )


    if korean_column == chinese_column:

        raise ColumnSelectionRequired(
            "韩文列和中文列不能相同。"
        )


    # =====================================================
    # 清掉旧报告
    #
    # 注意：
    # 这里只在内存工作簿中删除。
    # 原文件不会变化。
    # =====================================================

    remove_old_generated_sheets(
        workbook
    )


    header_map = build_header_map(
        worksheet=
            worksheet,

        header_row=
            header_row,
    )


    stats = (
        new_translation_stats()
    )


    details = []


    # =====================================================
    # 逐行处理
    # =====================================================

    total_rows = max(
        worksheet.max_row
        -
        header_row,
        0,
    )


    emit_progress(
        progress_callback=
            progress_callback,

        current=
            0,

        total=
            total_rows,

        stage=
            "翻译",

        message=
            "正在准备逐行翻译...",

        stats=
            stats,
    )


    for row_index in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):

        current = (
            row_index
            -
            header_row
            -
            1
        )


        emit_progress(
            progress_callback=
                progress_callback,

            current=
                current,

            total=
                total_rows,

            stage=
                "翻译",

            message=(
                f"正在处理第"
                f"{current + 1:,}"
                f"/"
                f"{total_rows:,}"
                f"行"
            ),

            row_index=
                row_index,

            stats=
                stats,
        )

        stats[
            "总数据行"
        ] += 1


        source_cell = worksheet.cell(
            row=row_index,
            column=korean_column,
        )


        target_cell = worksheet.cell(
            row=row_index,
            column=chinese_column,
        )


        # =================================================
        # 韩文源单元格是合并单元格结构
        # =================================================

        if isinstance(
            source_cell,
            MergedCell,
        ):

            stats[
                "合并单元格跳过"
            ] += 1


            details.append(
                {
                    "工作表":
                        worksheet.title,

                    "行号":
                        row_index,

                    "韩文":
                        "",

                    "中文":
                        "",

                    "处理方式":
                        "合并单元格跳过",

                    "来源":
                        "",

                    "格式检查":
                        "未执行",

                    "最终QA通过":
                        False,

                    "标点自动规范化":
                        False,

                    "标点规范化内容":
                        "",

                    "需人工确认":
                        True,

                    "确认原因":
                        "韩文源单元格属于合并单元格，禁止自动处理。",

                    "AI成功":
                        False,

                    "错误类型":
                        "",

                    "错误阶段":
                        "",

                    "错误":
                        "",
                }
            )


            continue


        # =================================================
        # 中文目标属于合并单元格
        # =================================================

        if isinstance(
            target_cell,
            MergedCell,
        ):

            stats[
                "合并单元格跳过"
            ] += 1


            details.append(
                {
                    "工作表":
                        worksheet.title,

                    "行号":
                        row_index,

                    "韩文":
                        cell_text(
                            source_cell
                        ),

                    "中文":
                        "",

                    "处理方式":
                        "合并单元格跳过",

                    "来源":
                        "",

                    "格式检查":
                        "未执行",

                    "最终QA通过":
                        False,

                    "标点自动规范化":
                        False,

                    "标点规范化内容":
                        "",

                    "需人工确认":
                        True,

                    "确认原因":
                        "中文目标单元格属于合并单元格，禁止自动写入。",

                    "AI成功":
                        False,

                    "错误类型":
                        "",

                    "错误阶段":
                        "",

                    "错误":
                        "",
                }
            )


            continue


        # =================================================
        # 公式保护
        # =================================================

        if (
            is_formula_cell(
                source_cell
            )
            or
            is_formula_cell(
                target_cell
            )
        ):

            stats[
                "公式跳过"
            ] += 1


            details.append(
                {
                    "工作表":
                        worksheet.title,

                    "行号":
                        row_index,

                    "韩文":
                        cell_text(
                            source_cell
                        ),

                    "中文":
                        cell_text(
                            target_cell
                        ),

                    "处理方式":
                        "公式保护跳过",

                    "来源":
                        "",

                    "格式检查":
                        "未执行",

                    "最终QA通过":
                        False,

                    "标点自动规范化":
                        False,

                    "标点规范化内容":
                        "",

                    "需人工确认":
                        True,

                    "确认原因":
                        "源或目标单元格为公式，系统禁止自动修改。",

                    "AI成功":
                        False,

                    "错误类型":
                        "",

                    "错误阶段":
                        "",

                    "错误":
                        "",
                }
            )


            continue


        korean_text = str(
            source_cell.value
            or
            ""
        )


        existing_chinese = str(
            target_cell.value
            or
            ""
        ).strip()


        # =================================================
        # 韩文空文本
        # =================================================

        if not korean_text:

            stats[
                "空文本跳过"
            ] += 1

            continue


        # =================================================
        # 非韩文
        # =================================================

        if not contains_korean(
            korean_text
        ):

            stats[
                "非韩文跳过"
            ] += 1

            continue


        # =================================================
        # 已有中文
        #
        # 默认绝不覆盖。
        # =================================================

        if existing_chinese:

            stats[
                "韩文文本数"
            ] += 1


            stats[
                "已有中文保留"
            ] += 1


            details.append(
                {
                    "工作表":
                        worksheet.title,

                    "行号":
                        row_index,

                    "韩文":
                        korean_text,

                    "中文":
                        existing_chinese,

                    "处理方式":
                        "已有中文保留",

                    "来源":
                        "原文件",

                    "格式检查":
                        "未执行",

                    "最终QA通过":
                        None,

                    "标点自动规范化":
                        False,

                    "标点规范化内容":
                        "",

                    "需人工确认":
                        False,

                    "确认原因":
                        "",

                    "AI成功":
                        False,

                    "错误类型":
                        "",

                    "错误阶段":
                        "",

                    "错误":
                        "",
                }
            )


            continue


        # =================================================
        # 如果是新建的AI中文列，
        # 复制韩文单元格样式。
        # =================================================

        if created_chinese_column:

            copy_cell_style(
                source_cell,
                target_cell,
            )


        # =================================================
        # Excel上下文
        # =================================================

        excel_context = (
            build_excel_context(
                worksheet=
                    worksheet,

                row_index=
                    row_index,

                header_row=
                    header_row,

                korean_column=
                    korean_column,

                header_map=
                    header_map,

                chinese_column=
                    chinese_column,
            )
        )


        # =================================================
        # 唯一统一翻译入口
        # =================================================

        result = (
            translate_text_unit(
                text=
                    korean_text,

                knowledge_base=
                    knowledge_base,

                extra_context=
                    excel_context,
            )
        )


        # =================================================
        # 致命API错误
        #
        # ★ 必须在写Excel之前判断 ★
        #
        # 如果失败：
        #
        # 立即raise
        # ↓
        # 整批停止
        # ↓
        # workbook不保存
        # ↓
        # 不产生部分文件
        # =================================================

        raise_if_fatal_ai_error(
            result=
                result,

            location=(
                f"Excel工作表"
                f"“{worksheet.title}”"
                f"第{row_index}行"
            ),
        )


        update_translation_stats(
            stats=
                stats,

            result=
                result,
        )


        # =================================================
        # AI非致命失败
        #
        # 不写韩文到中文目标列。
        # 保持空白。
        # =================================================

        if (
            result.get(
                "处理方式"
            )
            ==
            "AI调用失败，保留韩文"
        ):

            details.append(
                {
                    "工作表":
                        worksheet.title,

                    "行号":
                        row_index,

                    "韩文":
                        korean_text,

                    "中文":
                        "",

                    "处理方式":
                        result.get(
                            "处理方式",
                            "",
                        ),

                    "来源":
                        result.get(
                            "来源",
                            "",
                        ),

                    "格式检查":
                        result.get(
                            "格式检查",
                            "",
                        ),

                    "最终QA通过":
                        result.get(
                            "最终QA通过"
                        ),

                    "标点自动规范化":
                        result.get(
                            "标点自动规范化",
                            False,
                        ),

                    "标点规范化内容":
                        result.get(
                            "标点规范化内容",
                            "",
                        ),

                    "需人工确认":
                        True,

                    "确认原因":
                        result.get(
                            "确认原因",
                            "",
                        ),

                    "AI成功":
                        False,

                    "错误类型":
                        result.get(
                            "错误类型",
                            "",
                        ),

                    "错误阶段":
                        result.get(
                            "错误阶段",
                            "",
                        ),

                    "错误":
                        result.get(
                            "错误",
                            "",
                        ),

                    "耗时秒":
                        result.get(
                            "耗时秒"
                        ),

                    "模型":
                        result.get(
                            "模型",
                            "",
                        ),
                }
            )


            continue


        final_translation = str(
            result.get(
                "译文",
                ""
            )
            or
            ""
        )


        # =================================================
        # 防御性清理
        #
        # translate_text_unit理论上已经清理完。
        #
        # 如果未来某模块又带入marker，
        # 这里最后再拦一次。
        # =================================================

        (
            cleaned_translation,
            extra_manual_reason,
        ) = strip_manual_marker(
            final_translation
        )


        if cleaned_translation != final_translation:

            final_translation = (
                cleaned_translation
            )


        need_manual = bool(
            result.get(
                "需人工确认",
                False,
            )
        )


        confirmation_reason = str(
            result.get(
                "确认原因",
                ""
            )
            or
            ""
        )


        if extra_manual_reason:

            need_manual = True


            if confirmation_reason:

                confirmation_reason = (
                    confirmation_reason
                    +
                    "；"
                    +
                    extra_manual_reason
                )

            else:

                confirmation_reason = (
                    extra_manual_reason
                )


        # =================================================
        # 写入中文目标列
        #
        # 即便“需人工确认”，仍写入AI建议译文，
        # 但会进入需人工确认/格式异常报告。
        #
        # 原文件不会被覆盖。
        # =================================================

        target_cell.value = (
            final_translation
        )


        stats[
            "新增译文"
        ] += 1


        detail = {
            "工作表":
                worksheet.title,

            "行号":
                row_index,

            "韩文":
                korean_text,

            "中文":
                final_translation,

            "处理方式":
                result.get(
                    "处理方式",
                    "",
                ),

            "来源":
                result.get(
                    "来源",
                    "",
                ),

            "格式检查":
                result.get(
                    "格式检查",
                    "",
                ),

            "最终QA通过":
                result.get(
                    "最终QA通过"
                ),

            "标点自动规范化":
                result.get(
                    "标点自动规范化",
                    False,
                ),

            "标点规范化内容":
                result.get(
                    "标点规范化内容",
                    "",
                ),

            "需人工确认":
                need_manual,

            "确认原因":
                confirmation_reason,

            "AI成功":
                result.get(
                    "AI成功",
                    False,
                ),

            "错误类型":
                result.get(
                    "错误类型",
                    "",
                ),

            "错误阶段":
                result.get(
                    "错误阶段",
                    "",
                ),

            "错误":
                result.get(
                    "错误",
                    "",
                ),

            "耗时秒":
                result.get(
                    "耗时秒"
                ),

            "模型":
                result.get(
                    "模型",
                    "",
                ),
        }


        # =================================================
        # 常用元数据加入报告
        # =================================================

        for field_name in [
            "ID",
            "msgctxt",
            "references",
            "Release",
            "Content",
            "script",
        ]:

            value = (
                find_metadata_value(
                    worksheet=
                        worksheet,

                    row_index=
                        row_index,

                    header_map=
                        header_map,

                    aliases=
                        METADATA_ALIASES.get(
                            field_name,
                            [],
                        ),
                )
            )


            if value:

                detail[
                    field_name
                ] = value


        details.append(
            detail
        )


    # =====================================================
    # 输出配置
    # =====================================================

    config = {
        "处理工作表":
            worksheet.title,

        "表头行":
            header_row,

        "韩文列":
            (
                f"{get_column_letter(korean_column)}"
                f" / "
                f"{worksheet.cell(header_row, korean_column).value}"
            ),

        "中文列":
            (
                f"{get_column_letter(chinese_column)}"
                f" / "
                f"{worksheet.cell(header_row, chinese_column).value}"
            ),

        "中文列是否新建":
            (
                "是"
                if created_chinese_column
                else
                "否"
            ),

        "已有中文策略":
            "默认保留，不覆盖",

        "最终QA":
            (
                "translate_text_unit统一执行"
            ),

        "换行保护":
            "启用",

        "韩文残留检测":
            "启用",

        "安全标点自动规范化":
            "启用",

        "致命API错误策略":
            "第一次失败立即停止整批，不生成部分结果",
    }


    # =====================================================
    # 全部数据处理完成后
    # 才添加报告并保存。
    #
    # 这样可以保证：
    #
    # 致命API错误发生时
    # 不会生成半成品XLSX。
    # =====================================================

    emit_progress(
        progress_callback=
            progress_callback,

        current=
            total_rows,

        total=
            total_rows,

        stage=
            "生成报告",

        message=
            "翻译完成，正在生成审计报告并保存新Excel...",

        stats=
            stats,
    )


    add_translation_reports(
        workbook=
            workbook,

        stats=
            stats,

        details=
            details,

        config=
            config,
    )


    output = BytesIO()


    workbook.save(
        output
    )


    output.seek(
        0
    )


    emit_progress(
        progress_callback=
            progress_callback,

        current=
            total_rows,

        total=
            total_rows,

        stage=
            "完成",

        message=
            "XLSX智能翻译和报告生成全部完成。",

        stats=
            stats,
    )


    return (
        output.getvalue(),
        stats,
        details,
        config,
    )


# =========================================================
# 独立加载测试
# =========================================================

if __name__ == "__main__":

    print(
        "XLSX智能翻译模块加载成功。"
    )

    print(
        "已启用："
    )

    print(
        "1. translate_text_unit统一翻译链"
    )

    print(
        "2. 真实换行保护"
    )

    print(
        "3. 韩文残留最终QA"
    )

    print(
        "4. 安全中文标点自动规范化"
    )

    print(
        "5. 最终QA报告"
    )

    print(
        "6. 已有中文默认保留"
    )

    print(
        "7. 致命API错误整批安全停止"
    )
