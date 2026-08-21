from copy import copy

from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# =========================================================
# 系统生成的工作表
# =========================================================

GENERATED_SHEETS = [
    "待人工处理",
    "D级_冲突记录",
    "E级_完全未命中",
    "分析摘要",
]


# =========================================================
# 删除旧的系统生成工作表
# =========================================================

def remove_old_generated_sheets(workbook):
    """
    如果用户把之前生成过的报告再次上传，
    先删除旧的系统报告页，避免重复。
    """

    for sheet_name in GENERATED_SHEETS:

        if sheet_name in workbook.sheetnames:

            sheet = workbook[
                sheet_name
            ]

            workbook.remove(
                sheet
            )


# =========================================================
# 查找标题列
# =========================================================

def build_header_map(sheet):
    """
    返回：
    {
        "检索_可信等级": 10,
        "检索_匹配状态": 11,
        ...
    }
    """

    result = {}

    for column in range(
        1,
        sheet.max_column + 1
    ):

        value = sheet.cell(
            row=1,
            column=column
        ).value

        if value is None:
            continue

        result[
            str(value).strip()
        ] = column

    return result


# =========================================================
# 安全取值
# =========================================================

def get_value(
    sheet,
    row,
    header_map,
    header_name
):
    """
    根据标题名称安全读取单元格。
    """

    column = header_map.get(
        header_name
    )

    if column is None:
        return ""

    value = sheet.cell(
        row=row,
        column=column
    ).value

    if value is None:
        return ""

    return value


# =========================================================
# 报告表统一格式
# =========================================================

def format_report_sheet(sheet):
    """
    给系统生成的工作表做基础格式。
    """

    # 冻结标题
    sheet.freeze_panes = "A2"

    # 自动筛选
    if (
        sheet.max_row >= 1
        and
        sheet.max_column >= 1
    ):

        sheet.auto_filter.ref = (
            sheet.dimensions
        )

    # 标题
    for cell in sheet[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

    # 正文自动换行
    for row in sheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    # 列宽
    default_widths = {
        "A": 12,
        "B": 45,
        "C": 18,
        "D": 28,
        "E": 25,
        "F": 40,
        "G": 40,
        "H": 40,
        "I": 45,
        "J": 45,
        "K": 55,
        "L": 55,
        "M": 15,
        "N": 60,
    }

    for column_letter, width in (
        default_widths.items()
    ):

        if (
            sheet.max_column
            >=
            ord(column_letter)
            -
            ord("A")
            +
            1
        ):

            sheet.column_dimensions[
                column_letter
            ].width = width


# =========================================================
# 写标题
# =========================================================

REVIEW_HEADERS = [
    "原表行号",
    "韩文原文",
    "可信等级",
    "匹配状态",
    "参考知识库",
    "推荐中文",
    "UWO精确译文",
    "Quest精确译文",
    "UWO长术语",
    "Quest长术语",
    "UWO历史上下文",
    "Quest历史上下文",
    "需人工确认",
    "说明",
]


def write_review_header(sheet):
    """
    创建待处理报告统一标题。
    """

    for column, header in enumerate(
        REVIEW_HEADERS,
        start=1
    ):

        sheet.cell(
            row=1,
            column=column
        ).value = header


# =========================================================
# 从原表提取一行
# =========================================================

def build_review_row(
    source_sheet,
    source_row,
    korean_column,
    header_map
):
    """
    将原工作表某一行整理成
    待人工处理报告格式。
    """

    korean_value = (
        source_sheet.cell(
            row=source_row,
            column=korean_column
        ).value
    )

    if korean_value is None:
        korean_value = ""

    return [
        source_row,

        korean_value,

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_可信等级"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_匹配状态"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_参考知识库"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_推荐中文"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_UWO精确译文"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_Quest精确译文"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_UWO长术语"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_Quest长术语"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_UWO历史上下文"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_Quest历史上下文"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_需人工确认"
        ),

        get_value(
            source_sheet,
            source_row,
            header_map,
            "检索_说明"
        ),
    ]


# =========================================================
# 创建人工处理工作表
# =========================================================

def create_review_sheets(
    workbook,
    source_sheet_name,
    korean_column
):
    """
    自动创建：

    1. 待人工处理
    2. D级_冲突记录
    3. E级_完全未命中
    """

    remove_old_generated_sheets(
        workbook
    )

    source_sheet = workbook[
        source_sheet_name
    ]

    header_map = build_header_map(
        source_sheet
    )

    # =====================================================
    # 创建工作表
    # =====================================================

    review_sheet = workbook.create_sheet(
        "待人工处理"
    )

    conflict_sheet = workbook.create_sheet(
        "D级_冲突记录"
    )

    unmatched_sheet = workbook.create_sheet(
        "E级_完全未命中"
    )

    write_review_header(
        review_sheet
    )

    write_review_header(
        conflict_sheet
    )

    write_review_header(
        unmatched_sheet
    )


    # =====================================================
    # 统计
    # =====================================================

    summary = {
        "A级": 0,
        "B级": 0,
        "C级": 0,
        "D级": 0,
        "E级": 0,
        "角色名专项": 0,
        "待人工处理": 0,
        "D级冲突": 0,
        "E级未命中": 0,
    }


    # =====================================================
    # 扫描原工作表
    # =====================================================

    for row in range(
        2,
        source_sheet.max_row + 1
    ):

        confidence = str(
            get_value(
                source_sheet,
                row,
                header_map,
                "检索_可信等级"
            )
        ).strip()

        need_confirm = str(
            get_value(
                source_sheet,
                row,
                header_map,
                "检索_需人工确认"
            )
        ).strip()


        # ---------------------------------------------
        # 等级统计
        # ---------------------------------------------

        if confidence == "A级":

            summary[
                "A级"
            ] += 1

        elif confidence == "B级":

            summary[
                "B级"
            ] += 1

        elif confidence == "C级":

            summary[
                "C级"
            ] += 1

        elif confidence == "D级":

            summary[
                "D级"
            ] += 1

        elif confidence == "E级":

            summary[
                "E级"
            ] += 1

        elif confidence.startswith(
            "A级角色名"
        ):

            summary[
                "角色名专项"
            ] += 1


        # ---------------------------------------------
        # 是否进入人工处理清单
        # ---------------------------------------------

        manual_required = (
            need_confirm == "是"
            or
            confidence
            in [
                "B级",
                "C级",
                "D级",
                "E级",
            ]
            or
            confidence.startswith(
                "A级角色名"
            )
        )


        if manual_required:

            review_row = build_review_row(
                source_sheet,
                row,
                korean_column,
                header_map
            )

            review_sheet.append(
                review_row
            )

            summary[
                "待人工处理"
            ] += 1


        # ---------------------------------------------
        # D级
        # ---------------------------------------------

        if confidence == "D级":

            conflict_sheet.append(
                build_review_row(
                    source_sheet,
                    row,
                    korean_column,
                    header_map
                )
            )

            summary[
                "D级冲突"
            ] += 1


        # ---------------------------------------------
        # E级
        # ---------------------------------------------

        if confidence == "E级":

            unmatched_sheet.append(
                build_review_row(
                    source_sheet,
                    row,
                    korean_column,
                    header_map
                )
            )

            summary[
                "E级未命中"
            ] += 1


    # =====================================================
    # 格式
    # =====================================================

    format_report_sheet(
        review_sheet
    )

    format_report_sheet(
        conflict_sheet
    )

    format_report_sheet(
        unmatched_sheet
    )

    return summary


# =========================================================
# 分析摘要
# =========================================================

def create_summary_sheet(
    workbook,
    review_summary,
    batch_stats=None
):
    """
    创建适合公司查看的“分析摘要”。
    """

    if "分析摘要" in workbook.sheetnames:

        workbook.remove(
            workbook["分析摘要"]
        )

    sheet = workbook.create_sheet(
        "分析摘要",
        0
    )

    sheet["A1"] = (
        "韩中游戏本地化知识库分析摘要"
    )

    sheet["A1"].font = Font(
        bold=True,
        size=16
    )

    sheet.merge_cells(
        "A1:B1"
    )

    sheet["A3"] = "项目"
    sheet["B3"] = "数量"


    rows = [
        (
            "A级：完整句可直接复用",
            review_summary.get(
                "A级",
                0
            )
        ),
        (
            "B级：长术语参考",
            review_summary.get(
                "B级",
                0
            )
        ),
        (
            "C级：历史上下文参考",
            review_summary.get(
                "C级",
                0
            )
        ),
        (
            "D级：多译法/冲突",
            review_summary.get(
                "D级",
                0
            )
        ),
        (
            "E级：完全未命中",
            review_summary.get(
                "E级",
                0
            )
        ),
        (
            "正式角色名专项",
            review_summary.get(
                "角色名专项",
                0
            )
        ),
        (
            "待人工处理总数",
            review_summary.get(
                "待人工处理",
                0
            )
        ),
        (
            "D级冲突记录",
            review_summary.get(
                "D级冲突",
                0
            )
        ),
        (
            "E级完全未命中",
            review_summary.get(
                "E级未命中",
                0
            )
        ),
    ]


    current_row = 4

    for name, value in rows:

        sheet.cell(
            row=current_row,
            column=1
        ).value = name

        sheet.cell(
            row=current_row,
            column=2
        ).value = value

        current_row += 1


    # =====================================================
    # 如果有批量统计，再添加
    # =====================================================

    if batch_stats:

        current_row += 2

        sheet.cell(
            row=current_row,
            column=1
        ).value = "批量处理统计"

        sheet.cell(
            row=current_row,
            column=1
        ).font = Font(
            bold=True
        )

        current_row += 1

        for key, value in (
            batch_stats.items()
        ):

            sheet.cell(
                row=current_row,
                column=1
            ).value = key

            sheet.cell(
                row=current_row,
                column=2
            ).value = value

            current_row += 1


    # =====================================================
    # 说明
    # =====================================================

    current_row += 2

    sheet.cell(
        row=current_row,
        column=1
    ).value = "等级说明"

    sheet.cell(
        row=current_row,
        column=1
    ).font = Font(
        bold=True
    )

    current_row += 1


    descriptions = [
        (
            "A级",
            "存在唯一完整句精确历史译文，可作为高可信复用结果。"
        ),
        (
            "B级",
            "没有完整句，但发现正式主库中的长术语历史译法。"
        ),
        (
            "C级",
            "没有完整句或长术语，但发现相关历史上下文。"
        ),
        (
            "D级",
            "存在多个历史译法，或UWO与Quest译法冲突，需要人工判断。"
        ),
        (
            "E级",
            "三个正式知识库均没有足够可靠的历史参考。"
        ),
    ]


    for level, description in descriptions:

        sheet.cell(
            row=current_row,
            column=1
        ).value = level

        sheet.cell(
            row=current_row,
            column=2
        ).value = description

        current_row += 1


    # =====================================================
    # 格式
    # =====================================================

    sheet.column_dimensions[
        "A"
    ].width = 32

    sheet.column_dimensions[
        "B"
    ].width = 75

    for row in sheet.iter_rows():

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    return sheet


# =========================================================
# 总入口
# =========================================================

def add_review_reports(
    workbook,
    source_sheet_name,
    korean_column,
    batch_stats=None
):
    """
    batch_processor最终只需要调用这个函数。

    自动生成：

    - 待人工处理
    - D级_冲突记录
    - E级_完全未命中
    - 分析摘要
    """

    summary = create_review_sheets(
        workbook,
        source_sheet_name,
        korean_column
    )

    create_summary_sheet(
        workbook,
        summary,
        batch_stats
    )

    return summary